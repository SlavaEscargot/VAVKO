import sys
import os
import sqlite3
import tempfile
import shutil
import traceback
from datetime import datetime
from io import BytesIO

import pandas as pd
from PyQt6.QtWidgets import *
from PyQt6.QtCore import *
from PyQt6.QtGui import *
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog

from PIL import Image, ImageEnhance
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Проверка зависимостей
PIL_AVAILABLE = True
PANDAS_AVAILABLE = True
REPORTLAB_AVAILABLE = True

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Предупреждение: openpyxl не установлен. Экспорт в Excel недоступен.")
    print("Установите: pip install openpyxl")


class ModernDatabaseApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.image_columns = []
        self.photo_cache = {}
        self.db_name = None
        self.current_table = None
        self.connection = None
        self.joined_tables = []
        self.selected_attributes = []
        self.table_joins = {}
        self.image_references = []

        self.initUI()
        self.select_database_file()

    def initUI(self):
        """Инициализация пользовательского интерфейса"""
        self.setWindowTitle("SQLite3 Database Manager - Modern")
        self.setGeometry(100, 100, 1400, 900)

        # Центральный виджет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Главный layout
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Заголовок
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)

        title_label = QLabel("🗃️ SQLite Database Manager")
        title_font = QFont("Segoe UI", 16, QFont.Weight.Bold)
        title_label.setFont(title_font)

        hotkeys_label = QLabel("🔥 Горячие клавиши: F5=Обновить, Ctrl+S=Сохранить, Del=Удалить, Ctrl+P=Печать")
        hotkeys_label.setFont(QFont("Segoe UI", 8))

        self.db_label = QLabel("📁 База данных: не выбрана")

        header_layout.addWidget(title_label)
        header_layout.addWidget(hotkeys_label)
        header_layout.addStretch()
        header_layout.addWidget(self.db_label)

        # Панель быстрых действий
        quick_actions_group = QGroupBox("🚀 Быстрые действия")
        quick_actions_layout = QGridLayout()

        actions = [
            ("📊 Создать таблицу", self.create_table_dialog, "primary"),
            ("➕ Добавить запись", self.add_record_dialog, "success"),
            ("🗑️ Удалить таблицу", self.delete_table, "danger"),
            ("🔄 Обновить данные", self.refresh_data, "secondary"),
            ("🔗 Быстрое соединение", self.quick_join_tables, "primary"),
            ("👁️ Выбрать атрибуты", self.select_attributes_dialog, "secondary"),
            ("💾 Сменить БД", self.change_database, "secondary"),
            ("📝 Добавить колонку", self.add_column_dialog, "primary"),
            ("🖼️ Импорт Excel", self.import_excel, "success"),
            ("📤 Экспорт Excel", self.export_excel, "primary"),
            ("🖼️ Экспорт Excel с фото", self.export_excel_with_images_embedded, "success"),
            ("🖨️ Печать", self.print_data, "warning"),
            ("🔍 Исследовать БД", self.inspect_database, "primary"),
            ("🖼️ Найти все фото", self.find_and_display_all_photos, "success"),
            ("📷 Проверить фото", self.check_and_display_photos, "primary")
        ]

        row = 0
        col = 0
        for text, callback, style in actions:
            btn = QPushButton(text)
            btn.clicked.connect(callback)
            self.style_button(btn, style)
            quick_actions_layout.addWidget(btn, row, col)
            col += 1
            if col > 3:
                col = 0
                row += 1

        quick_actions_group.setLayout(quick_actions_layout)

        # Основной контент
        content_widget = QWidget()
        content_layout = QHBoxLayout(content_widget)

        # Левая панель
        left_panel = QWidget()
        left_panel.setFixedWidth(350)
        left_layout = QVBoxLayout(left_panel)

        # Список таблиц
        tables_group = QGroupBox("📋 Таблицы базы данных")
        tables_layout = QVBoxLayout()

        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("🔍 Поиск:"))
        self.table_search = QLineEdit()
        self.table_search.textChanged.connect(self.filter_tables)
        search_layout.addWidget(self.table_search)

        self.table_listbox = QListWidget()
        self.table_listbox.itemSelectionChanged.connect(self.on_table_select)

        tables_layout.addLayout(search_layout)
        tables_layout.addWidget(self.table_listbox)
        tables_group.setLayout(tables_layout)

        # Панель соединений
        joins_group = QGroupBox("🔗 Активные соединения")
        joins_layout = QVBoxLayout()

        self.join_info_text = QTextEdit()
        self.join_info_text.setReadOnly(True)
        self.join_info_text.setMaximumHeight(150)

        join_buttons_layout = QHBoxLayout()
        clear_joins_btn = QPushButton("🗑️ Очистить все")
        clear_joins_btn.clicked.connect(self.clear_joins)
        remove_join_btn = QPushButton("✂️ Удалить")
        remove_join_btn.clicked.connect(self.remove_join)
        advanced_join_btn = QPushButton("⚙️ Расширенное")
        advanced_join_btn.clicked.connect(self.join_tables_dialog)

        self.style_button(clear_joins_btn, "danger")
        self.style_button(remove_join_btn, "secondary")
        self.style_button(advanced_join_btn, "primary")

        join_buttons_layout.addWidget(clear_joins_btn)
        join_buttons_layout.addWidget(remove_join_btn)
        join_buttons_layout.addWidget(advanced_join_btn)

        joins_layout.addWidget(self.join_info_text)
        joins_layout.addLayout(join_buttons_layout)
        joins_group.setLayout(joins_layout)

        left_layout.addWidget(tables_group)
        left_layout.addWidget(joins_group)

        # Правая панель
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)

        # Панель инструментов данных
        data_tools_group = QGroupBox("📊 Данные таблицы")
        data_tools_layout = QVBoxLayout()

        # Сортировка
        sort_layout = QHBoxLayout()
        sort_layout.addWidget(QLabel("Сортировка:"))
        self.sort_column = QComboBox()
        self.sort_column.setFixedWidth(150)
        self.sort_order = QComboBox()
        self.sort_order.addItems(["По возрастанию", "По убыванию"])
        self.sort_order.setFixedWidth(150)
        apply_sort_btn = QPushButton("🔄 Применить")
        apply_sort_btn.clicked.connect(self.apply_sorting)

        sort_layout.addWidget(QLabel("По:"))
        sort_layout.addWidget(self.sort_column)
        sort_layout.addWidget(self.sort_order)
        sort_layout.addWidget(apply_sort_btn)
        sort_layout.addStretch()

        # Информация об атрибутах
        self.attributes_label = QLabel("👁️ Отображаемые атрибуты: все")

        # Кнопки редактирования
        edit_buttons_layout = QHBoxLayout()
        edit_btn = QPushButton("✏️ Редактировать")
        edit_btn.clicked.connect(self.edit_cell_value)
        delete_btn = QPushButton("🗑️ Удалить запись")
        delete_btn.clicked.connect(self.delete_record)
        rename_btn = QPushButton("📝 Переименовать атрибут")
        rename_btn.clicked.connect(self.rename_attribute_dialog)

        self.style_button(edit_btn, "primary")
        self.style_button(delete_btn, "danger")
        self.style_button(rename_btn, "secondary")

        edit_buttons_layout.addWidget(edit_btn)
        edit_buttons_layout.addWidget(delete_btn)
        edit_buttons_layout.addWidget(rename_btn)
        edit_buttons_layout.addStretch()

        data_tools_layout.addLayout(sort_layout)
        data_tools_layout.addWidget(self.attributes_label)
        data_tools_layout.addLayout(edit_buttons_layout)
        data_tools_group.setLayout(data_tools_layout)

        # Таблица данных
        self.table_widget = QTableWidget()
        self.table_widget.setAlternatingRowColors(True)
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_widget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_widget.customContextMenuRequested.connect(self.show_context_menu)
        self.table_widget.doubleClicked.connect(self.on_cell_double_click)

        right_layout.addWidget(data_tools_group)
        right_layout.addWidget(self.table_widget)

        content_layout.addWidget(left_panel)
        content_layout.addWidget(right_panel)

        # Статус бар
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("✅ Готов к работе")

        # Добавляем всё в главный layout
        main_layout.addWidget(header_widget)
        main_layout.addWidget(quick_actions_group)
        main_layout.addWidget(content_widget)

        # Настройка горячих клавиш
        self.setup_hotkeys()

    def style_button(self, button, style_type):
        """Стилизация кнопок"""
        if style_type == "primary":
            button.setStyleSheet("""
                QPushButton {
                    background-color: #007acc;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #005a9e;
                }
            """)
        elif style_type == "secondary":
            button.setStyleSheet("""
                QPushButton {
                    background-color: #6c757d;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #545b62;
                }
            """)
        elif style_type == "success":
            button.setStyleSheet("""
                QPushButton {
                    background-color: #28a745;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #218838;
                }
            """)
        elif style_type == "danger":
            button.setStyleSheet("""
                QPushButton {
                    background-color: #dc3545;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #c82333;
                }
            """)
        elif style_type == "warning":
            button.setStyleSheet("""
                QPushButton {
                    background-color: #ffc107;
                    color: #333333;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #e0a800;
                }
            """)

    def setup_hotkeys(self):
        """Настройка горячих клавиш"""
        # F5 - обновить
        refresh_shortcut = QShortcut(QKeySequence("F5"), self)
        refresh_shortcut.activated.connect(self.refresh_data)

        # Ctrl+S - сохранить
        save_shortcut = QShortcut(QKeySequence("Ctrl+S"), self)
        save_shortcut.activated.connect(self.quick_save)

        # Delete - удалить запись
        delete_shortcut = QShortcut(QKeySequence("Delete"), self)
        delete_shortcut.activated.connect(self.quick_delete)

        # Ctrl+P - печать
        print_shortcut = QShortcut(QKeySequence("Ctrl+P"), self)
        print_shortcut.activated.connect(self.print_data)

        # Enter - обновить/применить
        enter_shortcut = QShortcut(QKeySequence("Return"), self)
        enter_shortcut.activated.connect(self.on_enter_key)

    def on_enter_key(self):
        """Обработка клавиши Enter"""
        focused_widget = self.focusWidget()

        if isinstance(focused_widget, QLineEdit) or isinstance(focused_widget, QComboBox):
            # Если фокус в поле ввода, обновить данные
            self.refresh_data()
        elif isinstance(focused_widget, QTableWidget):
            # Если фокус в таблице, редактировать ячейку
            self.edit_cell_value()

    def quick_save(self):
        """Быстрое сохранение"""
        if self.connection:
            try:
                self.connection.commit()
                self.update_status("💾 Данные сохранены!")
            except sqlite3.Error as e:
                self.update_status(f"❌ Ошибка сохранения: {e}")

    def quick_delete(self):
        """Быстрое удаление"""
        if self.table_widget.selectionModel().hasSelection():
            self.delete_record()

    def select_database_file(self):
        """Выбор файла базы данных"""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Выберите файл базы данных",
            "",
            "SQLite Database (*.db);;All files (*.*)",
            "SQLite Database (*.db)"
        )

        if file_path:
            self.db_name = file_path
            if not file_path.endswith('.db'):
                self.db_name += '.db'
            self.connect_to_db()
        else:
            # Создаем базу по умолчанию
            self.db_name = "my_database.db"
            self.connect_to_db()

    def connect_to_db(self):
        """Подключение к базе данных"""
        try:
            self.connection = sqlite3.connect(self.db_name)
            self.connection.execute("PRAGMA foreign_keys = ON")
            self.update_table_list()
            self.update_db_label()
            self.update_status(f"✅ Подключено к базе данных: {os.path.basename(self.db_name)}")
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка подключения: {e}")

    def change_database(self):
        """Смена базы данных"""
        reply = QMessageBox.question(
            self,
            "Смена базы данных",
            "Вы уверены, что хотите сменить базу данных?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            if self.connection:
                self.connection.close()
            self.select_database_file()

    def update_table_list(self):
        """Обновление списка таблиц"""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = cursor.fetchall()

            self.table_listbox.clear()
            for table in tables:
                if table[0] != "sqlite_sequence":
                    self.table_listbox.addItem(table[0])
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка получения списка таблиц: {e}")

    def filter_tables(self):
        """Фильтрация таблиц"""
        search_term = self.table_search.text().lower()

        for i in range(self.table_listbox.count()):
            item = self.table_listbox.item(i)
            table_name = item.text()
            item.setHidden(search_term not in table_name.lower())

    def on_table_select(self):
        """Обработка выбора таблицы"""
        selected_items = self.table_listbox.selectedItems()
        if not selected_items:
            return

        new_table = selected_items[0].text()

        if self.current_table and self.joined_tables:
            self.table_joins[self.current_table] = self.joined_tables.copy()

        self.current_table = new_table
        self.joined_tables = self.table_joins.get(self.current_table, [])
        self.selected_attributes.clear()
        self.update_join_info()
        self.update_attributes_label()
        self.display_table_data()
        self.update_status(f"📊 Выбрана таблица: {new_table}")

    def delete_table(self):
        """Удаление таблицы"""
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Выберите таблицу для удаления!")
            return

        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Вы уверены, что хотите удалить таблицу '{self.current_table}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"DROP TABLE IF EXISTS {self.escape_table_name(self.current_table)}")
                self.connection.commit()

                self.update_status(f"✅ Таблица '{self.current_table}' удалена!")
                self.current_table = None
                self.joined_tables.clear()
                self.selected_attributes.clear()
                if self.current_table in self.table_joins:
                    del self.table_joins[self.current_table]
                self.update_table_list()
                self.clear_table()
                self.update_join_info()
                self.update_attributes_label()

            except sqlite3.Error as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка удаления таблицы: {e}")

    def display_table_data(self, sort_column=None, sort_order="ASC"):
        """Отображение данных таблицы"""
        if not self.current_table and not self.joined_tables:
            return

        try:
            self.table_widget.clear()
            query, display_columns = self.build_query(sort_column, sort_order)

            if not display_columns:
                QMessageBox.warning(self, "Предупреждение", "Нет атрибутов для отображения!")
                return

            cursor = self.connection.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()

            # Настраиваем таблицу
            self.table_widget.setRowCount(len(rows))
            self.table_widget.setColumnCount(len(display_columns))
            self.table_widget.setHorizontalHeaderLabels(display_columns)

            # Определяем колонки с фото
            self.image_columns = []
            for col in display_columns:
                if self.is_image_column(col):
                    self.image_columns.append(col)

            # Заполняем данными
            for row_idx, row in enumerate(rows):
                for col_idx, value in enumerate(row):
                    col_name = display_columns[col_idx]

                    if col_name in self.image_columns and value is not None and isinstance(value, bytes):
                        if self.is_valid_image_blob(value):
                            item = QTableWidgetItem("🖼️ Фото")
                            item.setData(Qt.ItemDataRole.UserRole, value)  # Сохраняем данные фото
                        else:
                            item = QTableWidgetItem("[BLOB данные]")
                    elif isinstance(value, bool):
                        item = QTableWidgetItem("✅ Да" if value else "❌ Нет")
                    elif value is None:
                        item = QTableWidgetItem("")
                    else:
                        item = QTableWidgetItem(str(value))

                    self.table_widget.setItem(row_idx, col_idx, item)

            # Настройка сортировки
            self.sort_column.clear()
            available_columns = self.get_available_columns()
            self.sort_column.addItems(available_columns)
            if available_columns:
                self.sort_column.setCurrentIndex(0)

            self.table_widget.resizeColumnsToContents()

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки данных: {e}")

    def clear_table(self):
        """Очистка таблицы"""
        self.table_widget.clear()
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(0)

    def build_query(self, sort_column=None, sort_order="ASC"):
        """Построение SQL запроса"""
        if not self.current_table:
            return "", []

        escaped_current_table = self.escape_table_name(self.current_table)
        used_columns = set()
        select_columns = []

        def add_columns(table_name):
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"PRAGMA table_info({self.escape_table_name(table_name)})")
                columns = cursor.fetchall()
                for col in columns:
                    col_name = col[1]
                    if col_name not in used_columns:
                        select_columns.append(
                            f"{self.escape_table_name(table_name)}.{self.escape_table_name(col_name)}")
                        used_columns.add(col_name)
            except sqlite3.Error:
                pass

        add_columns(self.current_table)
        for join_info in self.joined_tables:
            add_columns(join_info['table2'])

        if self.selected_attributes:
            final_columns = []
            used_columns.clear()
            for attr in self.selected_attributes:
                if '.' in attr:
                    table, col = attr.split('.')
                    if col not in used_columns:
                        final_columns.append(f"{self.escape_table_name(table)}.{self.escape_table_name(col)}")
                        used_columns.add(col)
                else:
                    if attr not in used_columns:
                        final_columns.append(self.escape_table_name(attr))
                        used_columns.add(attr)
            select_columns = final_columns

        if not select_columns:
            return "", []

        select_stmt = "SELECT " + ", ".join(select_columns)
        from_stmt = f"FROM {escaped_current_table}"

        join_stmts = []
        for join_info in self.joined_tables:
            join_type = join_info.get('join_type', 'INNER')
            table2 = self.escape_table_name(join_info['table2'])
            condition = join_info['condition']
            join_stmts.append(f"{join_type} JOIN {table2} ON {condition}")

        order_stmt = ""
        if sort_column:
            sql_order = "DESC" if sort_order == "По убыванию" else "ASC"
            order_stmt = f"ORDER BY {self.escape_table_name(sort_column)} {sql_order}"

        query = f"{select_stmt} {from_stmt} {' '.join(join_stmts)} {order_stmt}"

        display_columns = []
        for col in select_columns:
            clean_col = col.replace('"', '')
            if '.' in clean_col:
                display_columns.append(clean_col.split('.')[-1])
            else:
                display_columns.append(clean_col)

        return query.strip(), display_columns

    def is_image_column(self, column_name):
        """Проверка, является ли колонка колонкой с изображениями"""
        try:
            cursor = self.connection.cursor()

            # Проверяем основную таблицу
            if self.current_table:
                cursor.execute(f"PRAGMA table_info({self.escape_table_name(self.current_table)})")
                columns = cursor.fetchall()

                for col in columns:
                    if col[1] == column_name and col[2].upper() == 'BLOB':
                        return True

            # Проверяем соединенные таблицы
            for join_info in self.joined_tables:
                table_name = join_info['table2']
                try:
                    cursor.execute(f"PRAGMA table_info({self.escape_table_name(table_name)})")
                    columns = cursor.fetchall()

                    for col in columns:
                        if col[1] == column_name and col[2].upper() == 'BLOB':
                            return True
                except sqlite3.Error:
                    continue

            # Дополнительная проверка по имени колонки
            photo_keywords = ['photo', 'image', 'img', 'picture', 'pic', 'фото', 'изображение']
            if any(keyword in column_name.lower() for keyword in photo_keywords):
                return True

        except sqlite3.Error:
            pass

        return False

    def is_valid_image_blob(self, data):
        """Проверка валидности изображения"""
        if not isinstance(data, bytes):
            return False

        if len(data) < 100:
            return False

        try:
            # Проверяем магические числа форматов изображений
            if len(data) > 4:
                # JPEG: FF D8 FF
                if data[:3] == b'\xff\xd8\xff':
                    return True
                # PNG: 89 50 4E 47
                if data[:4] == b'\x89PNG':
                    return True
                # GIF: GIF87a или GIF89a
                if data[:6] in [b'GIF87a', b'GIF89a']:
                    return True
                # BMP: BM
                if data[:2] == b'BM':
                    return True
            return False
        except:
            return False

    def show_context_menu(self, position):
        """Показать контекстное меню"""
        menu = QMenu()

        copy_value_action = menu.addAction("📋 Копировать значение")
        copy_row_action = menu.addAction("📑 Копировать строку")
        copy_header_action = menu.addAction("🏷️ Копировать заголовок")
        menu.addSeparator()
        edit_action = menu.addAction("✏️ Редактировать значение")
        view_photo_action = menu.addAction("🖼️ Просмотреть фото")
        menu.addSeparator()
        delete_action = menu.addAction("🗑️ Удалить запись")

        action = menu.exec(self.table_widget.mapToGlobal(position))

        if action == copy_value_action:
            self.copy_cell_value()
        elif action == copy_row_action:
            self.copy_row()
        elif action == copy_header_action:
            self.copy_header()
        elif action == edit_action:
            self.edit_cell_value()
        elif action == view_photo_action:
            self.view_selected_image_full()
        elif action == delete_action:
            self.delete_record()

    def on_cell_double_click(self, index):
        """Обработка двойного клика по ячейке"""
        self.edit_cell_value()

    def edit_cell_value(self):
        """Редактирование значения ячейки"""
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            return

        row = selected_items[0].row()
        col = selected_items[0].column()
        column_name = self.table_widget.horizontalHeaderItem(col).text()
        current_value = self.table_widget.item(row, col).text()

        # Проверяем, не является ли это фото
        item_data = self.table_widget.item(row, col).data(Qt.ItemDataRole.UserRole)
        if item_data and isinstance(item_data, bytes):
            # Это фото, показываем диалог для фото
            self.add_photo_dialog(column_name, row, col)
            return

        table_name = self.get_column_table(column_name)
        if not table_name:
            QMessageBox.warning(self, "Ошибка", f"Не удалось определить таблицу для колонки '{column_name}'")
            return

        col_type = self.get_column_type(table_name, column_name)

        if col_type and col_type.upper() == 'BOOLEAN':
            dialog = BooleanEditDialog(self, column_name, current_value)
            if dialog.exec():
                new_value = dialog.get_value()
                self.update_cell_value(row, col, new_value, column_name, table_name)
        else:
            text, ok = QInputDialog.getText(
                self,
                f"Редактирование {column_name}",
                f"Введите новое значение для '{column_name}':",
                text=current_value
            )
            if ok and text != current_value:
                self.update_cell_value(row, col, text, column_name, table_name)

    def update_cell_value(self, row, col, new_value, column_name, table_name):
        """Обновление значения ячейки в базе данных"""
        try:
            cursor = self.connection.cursor()

            # Получаем первичный ключ
            cursor.execute(f"PRAGMA table_info({self.escape_table_name(table_name)})")
            columns_info = cursor.fetchall()
            primary_key_name = columns_info[0][1]

            # Получаем значение первичного ключа из отображаемых данных
            pk_col = -1
            for i in range(self.table_widget.columnCount()):
                if self.table_widget.horizontalHeaderItem(i).text() == primary_key_name:
                    pk_col = i
                    break

            if pk_col == -1:
                QMessageBox.critical(self, "Ошибка", "Не удалось определить первичный ключ!")
                return

            primary_key_value = self.table_widget.item(row, pk_col).text()

            # Обрабатываем значение в зависимости от типа
            processed_value = new_value
            col_type = self.get_column_type(table_name, column_name)
            if col_type and col_type.upper() == 'BOOLEAN':
                if new_value.lower() in ['true', '1', 'да', 'yes']:
                    processed_value = 1
                elif new_value.lower() in ['false', '0', 'нет', 'no']:
                    processed_value = 0

            query = f"UPDATE {self.escape_table_name(table_name)} SET {self.escape_table_name(column_name)} = ? WHERE {primary_key_name} = ?"
            cursor.execute(query, (processed_value, primary_key_value))
            self.connection.commit()

            # Обновляем отображение
            item = self.table_widget.item(row, col)
            if col_type and col_type.upper() == 'BOOLEAN':
                item.setText("✅ Да" if processed_value == 1 else "❌ Нет")
            else:
                item.setText(str(new_value))

            self.update_status(f"✅ Значение в таблице '{table_name}' обновлено!")

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка обновления значения: {e}")

    def add_photo_dialog(self, column_name, row, col):
        """Диалог добавления фото"""
        dialog = PhotoDialog(self, column_name)
        if dialog.exec():
            image_data = dialog.get_image_data()
            if image_data:
                self.update_image_value(row, col, image_data, column_name)

    def update_image_value(self, row, col, image_data, column_name):
        """Обновление значения изображения"""
        try:
            cursor = self.connection.cursor()

            # Получаем информацию о таблице
            cursor.execute(f"PRAGMA table_info({self.escape_table_name(self.current_table)})")
            columns_info = cursor.fetchall()
            primary_key_name = columns_info[0][1]

            # Получаем значение первичного ключа
            pk_col = -1
            for i in range(self.table_widget.columnCount()):
                if self.table_widget.horizontalHeaderItem(i).text() == primary_key_name:
                    pk_col = i
                    break

            if pk_col == -1:
                QMessageBox.critical(self, "Ошибка", "Не удалось определить первичный ключ!")
                return

            primary_key_value = self.table_widget.item(row, pk_col).text()

            query = f"UPDATE {self.escape_table_name(self.current_table)} SET {self.escape_table_name(column_name)} = ? WHERE {primary_key_name} = ?"
            cursor.execute(query, (image_data, primary_key_value))
            self.connection.commit()

            # Обновляем отображение
            item = self.table_widget.item(row, col)
            item.setText("🖼️ Фото")
            item.setData(Qt.ItemDataRole.UserRole, image_data)

            self.update_status("✅ Фото обновлено!")

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка обновления фото: {e}")

    def view_selected_image_full(self):
        """Просмотр полноразмерного фото"""
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            return

        row = selected_items[0].row()
        col = selected_items[0].column()
        column_name = self.table_widget.horizontalHeaderItem(col).text()

        if column_name not in self.image_columns:
            QMessageBox.warning(self, "Предупреждение", "Выбранная колонка не содержит фото!")
            return

        item = self.table_widget.item(row, col)
        image_data = item.data(Qt.ItemDataRole.UserRole)

        if not image_data or not isinstance(image_data, bytes):
            QMessageBox.warning(self, "Предупреждение", "В этой ячейке нет фото")
            return

        self.view_image(column_name, image_data)

    def view_image(self, column_name, image_data, record_info=""):
        """Просмотр изображения"""
        dialog = ImageViewDialog(self, column_name, image_data, record_info)
        dialog.exec()

    def copy_cell_value(self):
        """Копирование значения ячейки"""
        selected_items = self.table_widget.selectedItems()
        if selected_items:
            value = selected_items[0].text()
            QApplication.clipboard().setText(value)
            self.update_status("✅ Значение скопировано в буфер")

    def copy_row(self):
        """Копирование строки"""
        selected_items = self.table_widget.selectedItems()
        if selected_items:
            row = selected_items[0].row()
            row_data = []
            for col in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row, col)
                row_data.append(item.text() if item else "")

            row_text = "\t".join(row_data)
            QApplication.clipboard().setText(row_text)
            self.update_status("✅ Строка скопирована в буфер")

    def copy_header(self):
        """Копирование заголовка"""
        selected_items = self.table_widget.selectedItems()
        if selected_items:
            col = selected_items[0].column()
            header = self.table_widget.horizontalHeaderItem(col).text()
            QApplication.clipboard().setText(header)
            self.update_status("✅ Заголовок скопирован в буфер")

    def delete_record(self):
        """Удаление записи"""
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Предупреждение", "Выберите запись для удаления!")
            return

        reply = QMessageBox.question(
            self,
            "Подтверждение",
            "Вы уверены, что хотите удалить выбранную запись?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                cursor = self.connection.cursor()

                cursor.execute(f"PRAGMA table_info({self.escape_table_name(self.current_table)})")
                columns_info = cursor.fetchall()
                primary_key_name = columns_info[0][1]

                row = selected_items[0].row()
                pk_col = -1
                for i in range(self.table_widget.columnCount()):
                    if self.table_widget.horizontalHeaderItem(i).text() == primary_key_name:
                        pk_col = i
                        break

                if pk_col == -1:
                    QMessageBox.critical(self, "Ошибка", "Не удалось определить первичный ключ!")
                    return

                primary_key_value = self.table_widget.item(row, pk_col).text()

                query = f"DELETE FROM {self.escape_table_name(self.current_table)} WHERE {primary_key_name} = ?"
                cursor.execute(query, (primary_key_value,))
                self.connection.commit()

                self.table_widget.removeRow(row)
                self.update_status("✅ Запись удалена!")

            except sqlite3.Error as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка удаления записи: {e}")

    def rename_attribute_dialog(self):
        """Диалог переименования атрибута"""
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Сначала выберите таблицу!")
            return

        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape_table_name(self.current_table)})")
            columns = cursor.fetchall()

            if not columns:
                QMessageBox.warning(self, "Предупреждение", "В таблице нет атрибутов!")
                return

            column_names = [col[1] for col in columns]
            old_name, ok = QInputDialog.getItem(
                self,
                "Переименование атрибута",
                "Выберите атрибут для переименования:",
                column_names,
                0,
                False
            )

            if not ok or not old_name:
                return

            new_name, ok = QInputDialog.getText(
                self,
                "Переименование атрибута",
                f"Новое имя для атрибута '{old_name}':",
                text=old_name
            )

            if ok and new_name and new_name != old_name:
                self.rename_attribute(old_name, new_name)

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка получения структуры таблицы: {e}")

    def rename_attribute(self, old_name, new_name):
        """Переименование атрибута"""
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape_table_name(self.current_table)})")
            columns_info = cursor.fetchall()

            new_columns = []
            for col in columns_info:
                if col[1] == old_name:
                    new_columns.append(f'"{new_name}" {col[2]}')
                else:
                    new_columns.append(f'"{col[1]}" {col[2]}')

            temp_table = f"temp_{self.current_table}"
            create_query = f"CREATE TABLE {self.escape_table_name(temp_table)} ({', '.join(new_columns)})"
            cursor.execute(create_query)

            column_names = [f'"{col[1]}"' for col in columns_info]
            insert_query = f"INSERT INTO {self.escape_table_name(temp_table)} SELECT {', '.join(column_names)} FROM {self.escape_table_name(self.current_table)}"
            cursor.execute(insert_query)

            cursor.execute(f"DROP TABLE {self.escape_table_name(self.current_table)}")
            cursor.execute(
                f"ALTER TABLE {self.escape_table_name(temp_table)} RENAME TO {self.escape_table_name(self.current_table)}")

            self.connection.commit()
            self.display_table_data()
            self.update_status(f"✅ Атрибут '{old_name}' переименован в '{new_name}'!")

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка переименования атрибута: {e}")

    def add_column_dialog(self):
        """Диалог добавления колонки"""
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Сначала выберите таблицу!")
            return

        dialog = AddColumnDialog(self, self.current_table)
        if dialog.exec():
            column_name, column_type, default_value = dialog.get_data()
            self.add_column_to_table(column_name, column_type, default_value)

    def add_column_to_table(self, column_name, column_type, default_value=None):
        """Добавление колонки в таблицу"""
        try:
            cursor = self.connection.cursor()
            query = f"ALTER TABLE {self.escape_table_name(self.current_table)} ADD COLUMN {self.escape_table_name(column_name)} {column_type}"

            if default_value is not None:
                if column_type.upper() == 'BOOLEAN':
                    if default_value.lower() in ['true', '1', 'да', 'yes']:
                        default_value = '1'
                    else:
                        default_value = '0'
                query += f" DEFAULT {default_value}"

            cursor.execute(query)
            self.connection.commit()

            if default_value is not None:
                update_query = f"UPDATE {self.escape_table_name(self.current_table)} SET {self.escape_table_name(column_name)} = ?"
                cursor.execute(update_query, (default_value,))
                self.connection.commit()

            self.update_status(f"✅ Колонка '{column_name}' добавлена в таблицу '{self.current_table}'!")
            self.display_table_data()

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка добавления колонки: {e}")

    def get_column_table(self, column_name):
        """Определение таблицы для колонки"""
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape_table_name(self.current_table)})")
            columns = cursor.fetchall()
            for col in columns:
                if col[1] == column_name:
                    return self.current_table
        except sqlite3.Error:
            pass

        for join_info in self.joined_tables:
            table_name = join_info['table2']
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"PRAGMA table_info({self.escape_table_name(table_name)})")
                columns = cursor.fetchall()
                for col in columns:
                    if col[1] == column_name:
                        return table_name
            except sqlite3.Error:
                continue

        return None

    def get_column_type(self, table_name, column_name):
        """Получение типа колонки"""
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape_table_name(table_name)})")
            columns = cursor.fetchall()
            for col in columns:
                if col[1] == column_name:
                    return col[2]
        except sqlite3.Error:
            pass
        return None

    def get_available_columns(self):
        """Получение доступных колонок для сортировки"""
        columns_set = set()

        if self.current_table:
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"PRAGMA table_info({self.escape_table_name(self.current_table)})")
                table_columns = cursor.fetchall()
                for col in table_columns:
                    columns_set.add(col[1])
            except sqlite3.Error:
                pass

        for join_info in self.joined_tables:
            table_name = join_info['table2']
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"PRAGMA table_info({self.escape_table_name(table_name)})")
                table_columns = cursor.fetchall()
                for col in table_columns:
                    col_name = col[1]
                    if col_name not in columns_set:
                        columns_set.add(col_name)
            except sqlite3.Error:
                pass

        return sorted(list(columns_set))

    def get_all_tables_columns(self):
        """Получение всех колонок всех таблиц"""
        all_columns = {}
        used_columns = set()

        if self.current_table:
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"PRAGMA table_info({self.escape_table_name(self.current_table)})")
                columns = cursor.fetchall()
                table_columns = []
                for col in columns:
                    if col[1] not in used_columns:
                        table_columns.append(col[1])
                        used_columns.add(col[1])
                all_columns[self.current_table] = table_columns
            except sqlite3.Error:
                pass

        for join_info in self.joined_tables:
            table_name = join_info['table2']
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"PRAGMA table_info({self.escape_table_name(table_name)})")
                columns = cursor.fetchall()
                table_columns = []
                for col in columns:
                    if col[1] not in used_columns:
                        table_columns.append(col[1])
                        used_columns.add(col[1])
                all_columns[table_name] = table_columns
            except sqlite3.Error:
                pass

        return all_columns

    def update_attributes_label(self):
        """Обновление метки с атрибутами"""
        if self.selected_attributes:
            attrs_text = ", ".join([attr.split('.')[-1] for attr in self.selected_attributes[:3]])
            if len(self.selected_attributes) > 3:
                attrs_text += f"... (+{len(self.selected_attributes) - 3})"
            self.attributes_label.setText(f"👁️ Отображаемые атрибуты: {attrs_text}")
        else:
            self.attributes_label.setText("👁️ Отображаемые атрибуты: все")

    def apply_sorting(self):
        """Применение сортировки"""
        if (self.current_table or self.joined_tables) and self.sort_column.currentText():
            sort_order = self.sort_order.currentText()
            self.display_table_data(self.sort_column.currentText(), sort_order)

    def refresh_data(self):
        """Обновление данных"""
        if self.current_table or self.joined_tables:
            self.display_table_data()
        self.update_table_list()
        self.update_db_label()
        self.update_status("✅ Данные обновлены")

    def quick_join_tables(self):
        """Быстрое соединение таблиц"""
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Сначала выберите основную таблицу!")
            return

        tables = []
        for i in range(self.table_listbox.count()):
            table = self.table_listbox.item(i).text()
            if table != self.current_table:
                tables.append(table)

        if not tables:
            QMessageBox.information(self, "Информация", "Нет других таблиц для соединения!")
            return

        dialog = MultiTableSelectDialog(self, tables)
        if dialog.exec():
            selected_tables = dialog.get_selected_tables()
            for table2 in selected_tables:
                common_columns = self.find_common_columns(self.current_table, table2)

                if not common_columns:
                    QMessageBox.warning(
                        self,
                        "Предупреждение",
                        f"Не найдено общих полей между '{self.current_table}' и '{table2}'!"
                    )
                    continue

                join_column = common_columns[0]

                if self.join_tables(table2, join_column, join_column, "INNER"):
                    self.update_status(
                        f"✅ Автоматическое соединение: {self.current_table}.{join_column} = {table2}.{join_column}")

    def find_common_columns(self, table1, table2):
        """Поиск общих колонок"""
        try:
            cursor = self.connection.cursor()

            cursor.execute(f"PRAGMA table_info({self.escape_table_name(table1)})")
            table1_columns = [col[1] for col in cursor.fetchall()]

            cursor.execute(f"PRAGMA table_info({self.escape_table_name(table2)})")
            table2_columns = [col[1] for col in cursor.fetchall()]

            common_columns = list(set(table1_columns) & set(table2_columns))
            return common_columns

        except sqlite3.Error:
            return []

    def join_tables(self, table2, table1_attr, table2_attr, join_type="INNER"):
        """Соединение таблиц"""
        try:
            cursor = self.connection.cursor()

            cursor.execute(f"PRAGMA table_info({self.escape_table_name(self.current_table)})")
            table1_columns = [col[1] for col in cursor.fetchall()]
            if table1_attr not in table1_columns:
                QMessageBox.critical(self, "Ошибка", f"Атрибут '{table1_attr}' не найден!")
                return False

            cursor.execute(f"PRAGMA table_info({self.escape_table_name(table2)})")
            table2_columns = [col[1] for col in cursor.fetchall()]
            if table2_attr not in table2_columns:
                QMessageBox.critical(self, "Ошибка", f"Атрибут '{table2_attr}' не найден!")
                return False

            for join_info in self.joined_tables:
                if join_info['table2'] == table2:
                    QMessageBox.warning(self, "Предупреждение", f"Таблица '{table2}' уже соединена!")
                    return False

            condition = f"{self.escape_table_name(self.current_table)}.{self.escape_table_name(table1_attr)} = {self.escape_table_name(table2)}.{self.escape_table_name(table2_attr)}"

            join_info = {'table2': table2, 'condition': condition, 'join_type': join_type}
            self.joined_tables.append(join_info)
            self.table_joins[self.current_table] = self.joined_tables.copy()

            self.update_join_info()
            self.display_table_data()
            self.update_status(f"✅ Таблицы соединены: {self.current_table} ↔ {table2}")
            return True

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка соединения таблиц: {e}")
            return False

    def update_join_info(self):
        """Обновление информации о соединениях"""
        if self.joined_tables:
            text = f"Основная: {self.current_table}\n\n"
            for i, join_info in enumerate(self.joined_tables):
                text += f"{i + 1}. {join_info['table2']}\n"
                text += f"   Условие: {join_info['condition']}\n"
                text += f"   Тип: {join_info['join_type']}\n\n"
        else:
            text = "Нет активных соединений"

        self.join_info_text.setText(text)

    def remove_join(self):
        """Удаление соединения"""
        if not self.joined_tables:
            return

        if self.joined_tables:
            removed_join = self.joined_tables.pop()
            self.table_joins[self.current_table] = self.joined_tables.copy()
            self.update_join_info()
            self.display_table_data()
            self.update_status(f"✅ Соединение с '{removed_join['table2']}' удалено")

    def clear_joins(self):
        """Очистка всех соединений"""
        self.joined_tables.clear()
        if self.current_table:
            self.table_joins[self.current_table] = []
        self.update_join_info()
        if self.current_table:
            self.display_table_data()
        self.update_status("✅ Все соединения очищены")

    def print_data(self):
        """Печать данных в PDF"""
        if not self.current_table and not self.joined_tables:
            QMessageBox.warning(self, "Предупреждение", "Нет данных для печати!")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить PDF",
            "",
            "PDF files (*.pdf);;All files (*.*)",
            "PDF files (*.pdf)"
        )

        if not file_path:
            return

        try:
            # Получаем данные
            query, display_columns = self.build_query()
            cursor = self.connection.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()

            if not rows:
                QMessageBox.information(self, "Информация", "Нет данных для печати")
                return

            # Создаем PDF
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.pdfgen import canvas

            # Используем альбомную ориентацию
            pdf = canvas.Canvas(file_path, pagesize=landscape(A4))
            pdf.setTitle(f"База данных - {self.current_table}")

            # Настройка шрифта
            try:
                font_paths = [
                    "C:/Windows/Fonts/arial.ttf",
                    "C:/Windows/Fonts/arialbd.ttf",
                    "/usr/share/fonts/truetype/msttcorefonts/arial.ttf",
                ]

                for font_path in font_paths:
                    if os.path.exists(font_path):
                        pdfmetrics.registerFont(TTFont('Arial', font_path))
                        pdf.setFont('Arial', 12)
                        break
                else:
                    pdf.setFont("Helvetica", 12)
            except:
                pdf.setFont("Helvetica", 12)

            # Заголовок
            title = f"Таблица: {self.current_table}"
            pdf.setFontSize(16)
            pdf.drawString(50, 520, title)

            pdf.setFontSize(10)
            pdf.drawString(50, 500, f"База данных: {os.path.basename(self.db_name)}")
            pdf.drawString(50, 485, f"Дата экспорта: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}")

            # Настройки таблицы
            col_width = 120
            row_height = 100
            start_x = 50
            start_y = 450

            # Определяем колонки с фото
            image_columns = []
            for col in display_columns:
                if self.is_image_column(col):
                    image_columns.append(col)

            # Заголовки колонок
            pdf.setFontSize(8)
            for i, col in enumerate(display_columns):
                x = start_x + i * col_width
                pdf.rect(x, start_y, col_width, 20)
                safe_text = self.safe_text_for_pdf(str(col)[:15])
                pdf.drawString(x + 2, start_y + 5, safe_text)

            # Данные
            pdf.setFontSize(7)
            y_pos = start_y - 20
            temp_files = []

            for row_idx, row in enumerate(rows):
                if y_pos < 50:
                    pdf.showPage()
                    pdf.setFontSize(16)
                    pdf.drawString(50, 520, f"Таблица: {self.current_table} - продолжение")

                    y_pos = 450
                    pdf.setFontSize(8)
                    for i, col in enumerate(display_columns):
                        x = start_x + i * col_width
                        pdf.rect(x, y_pos, col_width, 20)
                        safe_text = self.safe_text_for_pdf(str(col)[:15])
                        pdf.drawString(x + 2, y_pos + 5, safe_text)
                    y_pos = y_pos - 20
                    pdf.setFontSize(7)

                for i, value in enumerate(row):
                    col_name = display_columns[i]
                    x = start_x + i * col_width

                    pdf.rect(x, y_pos, col_width, row_height)

                    if col_name in image_columns and value is not None and isinstance(value, bytes):
                        try:
                            if self.is_valid_image_blob(value):
                                with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                                    tmp.write(value)
                                    temp_file = tmp.name
                                    temp_files.append(temp_file)

                                try:
                                    image = Image.open(BytesIO(value))
                                    max_width = col_width - 4
                                    max_height = row_height - 4

                                    if image.width > max_width or image.height > max_height:
                                        ratio = min(max_width / image.width, max_height / image.height)
                                        new_size = (int(image.width * ratio), int(image.height * ratio))
                                        image = image.resize(new_size, Image.Resampling.LANCZOS)

                                    image.save(temp_file, format='PNG')
                                    img = ImageReader(temp_file)
                                    pdf.drawImage(img, x + 2, y_pos + 2,
                                                  width=max_width,
                                                  height=max_height,
                                                  preserveAspectRatio=True,
                                                  mask='auto')
                                except Exception as img_error:
                                    pdf.drawString(x + 2, y_pos + 40, "Изображение")
                                    pdf.drawString(x + 2, y_pos + 30, f"{len(value)} байт")
                            else:
                                pdf.drawString(x + 2, y_pos + 40, "Невалидное")
                                pdf.drawString(x + 2, y_pos + 30, f"{len(value)} байт")
                        except Exception as e:
                            pdf.drawString(x + 2, y_pos + 40, "Ошибка")
                            pdf.drawString(x + 2, y_pos + 30, str(e)[:20])
                    elif value is None:
                        pdf.drawString(x + 2, y_pos + 40, "")
                    elif isinstance(value, bool):
                        pdf.drawString(x + 2, y_pos + 40, "Да" if value else "Нет")
                    elif isinstance(value, (int, float)):
                        pdf.drawString(x + 2, y_pos + 40, str(value))
                    else:
                        text = str(value)
                        if len(text) > 20:
                            text = text[:17] + "..."
                        pdf.drawString(x + 2, y_pos + 40, text)

                y_pos -= row_height

                if y_pos < 50:
                    pdf.showPage()
                    pdf.setFontSize(16)
                    pdf.drawString(50, 520, f"Таблица: {self.current_table} - продолжение")

                    y_pos = 450
                    pdf.setFontSize(8)
                    for i, col in enumerate(display_columns):
                        x = start_x + i * col_width
                        pdf.rect(x, y_pos, col_width, 20)
                        safe_text = self.safe_text_for_pdf(str(col)[:15])
                        pdf.drawString(x + 2, y_pos + 5, safe_text)
                    y_pos = y_pos - 20
                    pdf.setFontSize(7)

            pdf.save()

            # Очищаем временные файлы
            for temp_file in temp_files:
                try:
                    os.unlink(temp_file)
                except:
                    pass

            self.update_status(f"✅ PDF создан: {os.path.basename(file_path)}")
            QMessageBox.information(self, "Успех", f"PDF успешно создан:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка создания PDF: {e}")
            if 'temp_files' in locals():
                for temp_file in temp_files:
                    try:
                        os.unlink(temp_file)
                    except:
                        pass

    def safe_text_for_pdf(self, text):
        """Безопасный текст для PDF"""
        if not text:
            return ""

        import re
        text = re.sub(r'[^\x20-\x7E\u0400-\u04FF]', '', text)

        if len(text) > 30:
            text = text[:27] + "..."

        return text

    def import_excel(self):
        """Импорт из Excel"""
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Сначала выберите таблицу!")
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите Excel файл",
            "",
            "Excel files (*.xlsx *.xls);;All files (*.*)"
        )

        if not file_path:
            return

        try:
            df = pd.read_excel(file_path)

            if df.empty:
                QMessageBox.warning(self, "Предупреждение", "Файл Excel пуст!")
                return

            dialog = ExcelImportDialog(self, df.columns.tolist())
            if not dialog.exec():
                return

            cursor = self.connection.cursor()

            cursor.execute(f"PRAGMA table_info({self.escape_table_name(self.current_table)})")
            table_columns = [col[1] for col in cursor.fetchall()]

            for _, row in df.iterrows():
                values = []
                for table_col in table_columns:
                    if table_col in df.columns:
                        value = row[table_col]
                        if pd.isna(value):
                            values.append(None)
                        else:
                            values.append(value)
                    else:
                        values.append(None)

                placeholders = ", ".join(["?" for _ in table_columns])
                query = f"INSERT INTO {self.escape_table_name(self.current_table)} VALUES ({placeholders})"
                cursor.execute(query, values)

            self.connection.commit()
            self.display_table_data()
            self.update_status(f"✅ Данные импортированы из {os.path.basename(file_path)}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка импорта Excel: {e}")

    def export_excel(self):
        """Экспорт в Excel (базовый)"""
        if not self.current_table and not self.joined_tables:
            QMessageBox.warning(self, "Предупреждение", "Нет данных для экспорта!")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить как Excel (базовый)",
            "",
            "Excel files (*.xlsx);;All files (*.*)",
            "Excel files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            query, display_columns = self.build_query()
            cursor = self.connection.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()

            df = pd.DataFrame(rows, columns=display_columns)

            for i, col in enumerate(display_columns):
                if self.is_image_column(col):
                    df[col] = ["🖼️ Фото" if isinstance(val, bytes) and self.is_valid_image_blob(val) else val for val in
                               df[col]]

            directory = os.path.dirname(file_path)
            if directory and not os.path.exists(directory):
                os.makedirs(directory)

            df.to_excel(file_path, index=False, engine='openpyxl')

            self.update_status(f"✅ Данные экспортированы в {os.path.basename(file_path)}")
            QMessageBox.information(self, "Успех", f"Данные успешно экспортированы в:\n{file_path}")

        except PermissionError as e:
            QMessageBox.critical(
                self,
                "Ошибка доступа",
                f"Нет прав доступа к файлу:\n{file_path}\n\n"
                f"Сохраните файл в другую папку (например, Документы или Рабочий стол)"
            )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта в Excel: {e}")

    def escape_table_name(self, table_name):
        """Экранирование имени таблицы"""
        return f'"{table_name}"'

    def update_db_label(self):
        """Обновление метки с именем БД"""
        if self.db_name:
            db_name = os.path.basename(self.db_name)
            self.db_label.setText(f"📁 База данных: {db_name}")

    def create_table_dialog(self):
        """Диалог создания таблицы"""
        dialog = CreateTableDialog(self)
        if dialog.exec():
            table_name, columns = dialog.get_data()
            self.create_table(table_name, columns)

    def create_table(self, table_name, columns):
        """Создание таблицы"""
        try:
            cursor = self.connection.cursor()
            columns_sql = []
            for col in columns:
                col_name = f'"{col["name"]}"'
                columns_sql.append(f"{col_name} {col['type']}")

            query = f"CREATE TABLE IF NOT EXISTS {self.escape_table_name(table_name)} ({', '.join(columns_sql)})"
            cursor.execute(query)
            self.connection.commit()

            self.update_status(f"✅ Таблица '{table_name}' создана успешно!")
            self.update_table_list()

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка создания таблицы: {e}")

    def add_record_dialog(self):
        """Диалог добавления записи"""
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Выберите таблицу!")
            return

        dialog = AddRecordDialog(self, self.current_table, self.connection)
        if dialog.exec():
            values = dialog.get_values()
            self.add_record(values)

    def add_record(self, values):
        """Добавление записи"""
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape_table_name(self.current_table)})")
            columns_info = cursor.fetchall()
            columns = [column[1] for column in columns_info]
            columns_types = [column[2] for column in columns_info]

            processed_values = []
            for i, value in enumerate(values):
                col_type = columns_types[i].upper()

                if value is None or value == "":
                    processed_values.append(None)
                elif col_type == 'BOOLEAN':
                    if isinstance(value, str):
                        value_lower = value.lower().strip()
                        if value_lower in ['true', '1', 'да', 'yes', 'истина']:
                            processed_values.append(1)
                        elif value_lower in ['false', '0', 'нет', 'no', 'ложь']:
                            processed_values.append(0)
                        else:
                            processed_values.append(None)
                    else:
                        processed_values.append(1 if value else 0)
                else:
                    processed_values.append(value)

            placeholders = ", ".join(["?" for _ in columns])
            query = f"INSERT INTO {self.escape_table_name(self.current_table)} VALUES ({placeholders})"

            cursor.execute(query, processed_values)
            self.connection.commit()

            self.update_status("✅ Запись добавлена успешно!")
            self.display_table_data()

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка добавления записи: {e}")

    def join_tables_dialog(self):
        """Диалог соединения таблиц"""
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Сначала выберите основную таблицу!")
            return

        dialog = JoinTablesDialog(self, self.current_table, self.connection)
        if dialog.exec():
            table2, attr1, attr2, join_type = dialog.get_data()
            self.join_tables(table2, attr1, attr2, join_type)

    def select_attributes_dialog(self):
        """Диалог выбора атрибутов"""
        if not self.current_table and not self.joined_tables:
            QMessageBox.warning(self, "Предупреждение", "Сначала выберите таблицу!")
            return

        dialog = SelectAttributesDialog(self, self.get_all_tables_columns(), self.selected_attributes)
        if dialog.exec():
            self.selected_attributes = dialog.get_selected_attributes()
            self.update_attributes_label()
            self.display_table_data()

    def check_and_display_photos(self):
        """Проверка наличия фото"""
        if not self.current_table:
            return

        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape_table_name(self.current_table)})")
            columns = cursor.fetchall()

            image_columns = []
            for col in columns:
                if col[2].upper() == 'BLOB':
                    image_columns.append(col[1])

            if image_columns:
                photo_found = False
                for col_name in image_columns:
                    cursor.execute(f"SELECT COUNT(*) FROM {self.current_table} WHERE {col_name} IS NOT NULL")
                    result = cursor.fetchone()

                    if result and result[0] > 0:
                        self.update_status(f"✅ Найдено {result[0]} фото в колонке '{col_name}'")
                        photo_found = True

                if not photo_found:
                    self.update_status("ℹ️ В таблице есть колонки для фото, но фото не найдены")

        except Exception as e:
            pass

    def inspect_database(self):
        """Исследование базы данных"""
        try:
            if not self.connection:
                QMessageBox.warning(self, "Предупреждение", "База данных не подключена!")
                return

            cursor = self.connection.cursor()

            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = cursor.fetchall()

            result_text = "🔍 ИССЛЕДОВАНИЕ БАЗЫ ДАННЫХ\n"
            result_text += "=" * 50 + "\n\n"
            result_text += f"📁 База данных: {os.path.basename(self.db_name)}\n"
            result_text += f"📋 Найдено таблиц: {len(tables)}\n\n"

            for table in tables:
                table_name = table[0]
                result_text += f"📊 ТАБЛИЦА: {table_name}\n"
                result_text += "-" * 30 + "\n"

                cursor.execute(f"PRAGMA table_info({self.escape_table_name(table_name)})")
                columns = cursor.fetchall()
                result_text += "Столбцы:\n"
                for col in columns:
                    result_text += f"  - {col[1]} (тип: {col[2]})\n"

                try:
                    cursor.execute(f"SELECT COUNT(*) FROM {self.escape_table_name(table_name)}")
                    count = cursor.fetchone()[0]
                    result_text += f"📈 Записей: {count}\n"
                except:
                    result_text += "📈 Записей: недоступно\n"

                result_text += "\n"

            self.show_text_dialog("Исследование базы данных", result_text)

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка исследования базы данных: {e}")

    def find_and_display_all_photos(self):
        """Поиск всех фотографий"""
        try:
            if not self.connection:
                QMessageBox.warning(self, "Предупреждение", "База данных не подключена!")
                return

            cursor = self.connection.cursor()

            tables = cursor.execute("SELECT name FROM sqlite_master WHERE type='table';").fetchall()

            photo_count = 0
            result_text = "🖼️ ПОИСК ФОТОГРАФИЙ В БАЗЕ ДАННЫХ\n"
            result_text += "=" * 50 + "\n\n"

            for table in tables:
                table_name = table[0]
                result_text += f"📋 Таблица: {table_name}\n"

                cursor.execute(f"PRAGMA table_info({self.escape_table_name(table_name)})")
                columns = cursor.fetchall()

                table_photo_count = 0
                for column in columns:
                    col_name = column[1]
                    col_type = column[2]

                    if (col_type.upper() == 'BLOB' or
                            any(photo_keyword in col_name.lower() for photo_keyword in
                                ['photo', 'image', 'img', 'picture', 'pic'])):

                        result_text += f"  🔍 Проверка столбца: {col_name} ({col_type})\n"

                        cursor.execute(f"SELECT rowid, {col_name} FROM {table_name} WHERE {col_name} IS NOT NULL")
                        photos = cursor.fetchall()

                        for rowid, photo_data in photos:
                            if isinstance(photo_data, bytes) and len(photo_data) > 100:
                                filename = f"photo_{table_name}_{col_name}_{rowid}.jpg"
                                try:
                                    with open(filename, 'wb') as f:
                                        f.write(photo_data)
                                    result_text += f"    ✅ Сохранено: {filename} ({len(photo_data)} bytes)\n"
                                    photo_count += 1
                                    table_photo_count += 1
                                except Exception as e:
                                    result_text += f"    ❌ Ошибка сохранения: {e}\n"
                            elif isinstance(photo_data, bytes):
                                result_text += f"    ℹ Найдены бинарные данные, но размер слишком мал для фото: {len(photo_data)} bytes\n"

                if table_photo_count == 0:
                    result_text += "  ❌ Фотографии не найдены\n"
                else:
                    result_text += f"  📊 Найдено фотографий: {table_photo_count}\n"

                result_text += "\n"

            if photo_count == 0:
                result_text += "⚠ Фотографии не найдены в базе данных\n"
            else:
                result_text += f"✅ Всего сохранено фотографий: {photo_count}\n"

            self.show_text_dialog("Результаты поиска фотографий", result_text)

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при поиске фотографий: {e}")

    def show_text_dialog(self, title, text):
        """Показать текстовый диалог"""
        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.setGeometry(100, 100, 800, 600)

        layout = QVBoxLayout(dialog)

        text_edit = QTextEdit()
        text_edit.setPlainText(text)
        text_edit.setReadOnly(True)
        text_edit.setFont(QFont("Consolas", 10))

        buttons_layout = QHBoxLayout()
        save_btn = QPushButton("💾 Сохранить в файл")
        save_btn.clicked.connect(lambda: self.save_text_to_file(text, title))
        close_btn = QPushButton("❌ Закрыть")
        close_btn.clicked.connect(dialog.close)

        buttons_layout.addWidget(save_btn)
        buttons_layout.addStretch()
        buttons_layout.addWidget(close_btn)

        layout.addWidget(text_edit)
        layout.addLayout(buttons_layout)

        dialog.exec()

    def save_text_to_file(self, text, title):
        """Сохранение текста в файл"""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            f"Сохранить {title}",
            "",
            "Text files (*.txt);;All files (*.*)",
            "Text files (*.txt)"
        )

        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(text)
                self.update_status(f"✅ Файл сохранен: {os.path.basename(file_path)}")
                QMessageBox.information(self, "Успех", f"Файл успешно сохранен:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка сохранения файла: {e}")

    def export_excel_with_images_embedded(self):
        """Экспорт в Excel с фото"""
        if not self.current_table and not self.joined_tables:
            QMessageBox.warning(self, "Предупреждение", "Нет данных для экспорта!")
            return

        dialog = ExportSettingsDialog(self)
        if not dialog.exec():
            return

        settings = dialog.get_settings()

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить как Excel",
            "",
            "Excel files (*.xlsx);;All files (*.*)",
            "Excel files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            query, display_columns = self.build_query()
            cursor = self.connection.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()

            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as ExcelImage
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = self.current_table or "Данные"

            # Записываем заголовки
            for col_idx, col_name in enumerate(display_columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 15

            photo_count = 0
            saved_files = []
            temp_dir = tempfile.mkdtemp(prefix="excel_export_")
            temp_files = []

            try:
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, value in enumerate(row, 1):
                        col_name = display_columns[col_idx - 1]

                        if (col_name in self.image_columns and
                                value is not None and
                                isinstance(value, bytes) and
                                settings['include_images']):

                            try:
                                if self.is_valid_image_blob(value):
                                    temp_file = os.path.join(temp_dir, f"photo_{row_idx}_{col_idx}.png")

                                    with open(temp_file, 'wb') as f:
                                        f.write(value)
                                    temp_files.append(temp_file)

                                    if settings['save_as_files']:
                                        save_dir = os.path.dirname(file_path) or "."
                                        photo_filename = f"{self.current_table}_row{row_idx - 1}_{col_name}.png"
                                        photo_path = os.path.join(save_dir, photo_filename)

                                        os.makedirs(save_dir, exist_ok=True)
                                        shutil.copy2(temp_file, photo_path)
                                        saved_files.append(photo_path)
                                        ws.cell(row=row_idx, column=col_idx, value=f"📷 {photo_filename}")
                                    else:
                                        try:
                                            img = ExcelImage(temp_file)
                                            img_size = settings['image_size']
                                            img.width = img_size
                                            img.height = img_size

                                            cell_coord = f"{get_column_letter(col_idx)}{row_idx}"
                                            ws.add_image(img, cell_coord)
                                            ws.row_dimensions[row_idx].height = img_size * 0.75
                                            photo_count += 1
                                        except Exception as img_error:
                                            ws.cell(row=row_idx, column=col_idx,
                                                    value=f"[Ошибка: {str(img_error)[:30]}]")

                                else:
                                    ws.cell(row=row_idx, column=col_idx, value="[Невалидное изображение]")

                            except Exception as e:
                                ws.cell(row=row_idx, column=col_idx, value=f"[Ошибка: {str(e)[:30]}]")

                        elif col_name in self.image_columns and value is not None:
                            ws.cell(row=row_idx, column=col_idx, value="🖼️ Фото")

                        elif isinstance(value, bool):
                            ws.cell(row=row_idx, column=col_idx, value="✅ Да" if value else "❌ Нет")

                        elif value is None:
                            ws.cell(row=row_idx, column=col_idx, value="")

                        else:
                            ws.cell(row=row_idx, column=col_idx, value=str(value))

                # Создаем лист с информацией
                ws_info = wb.create_sheet(title="Информация")
                ws_info['A1'] = "Отчет об экспорте"
                ws_info['A3'] = f"Таблица: {self.current_table}"
                ws_info['A4'] = f"Файл базы данных: {os.path.basename(self.db_name)}"
                ws_info['A5'] = f"Дата экспорта: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws_info['A6'] = f"Всего строк: {len(rows)}"
                ws_info['A7'] = f"Всего колонок: {len(display_columns)}"
                ws_info['A8'] = f"Фото в экспорте: {photo_count}"

                if saved_files:
                    ws_info['A10'] = "Сохраненные файлы фото:"
                    for i, file_path_saved in enumerate(saved_files, start=11):
                        ws_info[f'A{i}'] = os.path.basename(file_path_saved)

                wb.save(file_path)

                report = f"✅ Экспорт завершен успешно!\n\n"
                report += f"Файл: {os.path.basename(file_path)}\n"
                report += f"Расположение: {os.path.dirname(file_path)}\n"
                report += f"Строк данных: {len(rows)}\n"
                report += f"Колонок: {len(display_columns)}\n"

                if settings['include_images']:
                    if settings['save_as_files']:
                        report += f"Фото сохранены как файлы: {len(saved_files)}\n"
                    else:
                        report += f"Фото встроены в Excel: {photo_count}\n"

                self.update_status(f"✅ Экспорт завершен: {os.path.basename(file_path)}")
                QMessageBox.information(self, "Успешный экспорт", report)

            finally:
                for temp_file in temp_files:
                    try:
                        if os.path.exists(temp_file):
                            os.unlink(temp_file)
                    except:
                        pass

                try:
                    if os.path.exists(temp_dir):
                        os.rmdir(temp_dir)
                except:
                    pass

        except PermissionError as e:
            QMessageBox.critical(
                self,
                "Ошибка доступа",
                f"Нет прав доступа к файлу:\n{file_path}\n\n"
                f"Сохраните файл в другую папку (например, Документы или Рабочий стол)"
            )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта: {str(e)}")

    def update_status(self, message):
        """Обновление статуса"""
        self.status_bar.showMessage(message)
        QTimer.singleShot(3000, lambda: self.status_bar.showMessage("✅ Готов к работе"))


# ВСПОМОГАТЕЛЬНЫЕ ДИАЛОГИ

class BooleanEditDialog(QDialog):
    def __init__(self, parent, column_name, current_value):
        super().__init__(parent)
        self.setWindowTitle(f"Редактирование {column_name}")
        self.setGeometry(300, 300, 300, 150)

        layout = QVBoxLayout(self)

        label = QLabel(f"Выберите значение для '{column_name}':")
        layout.addWidget(label)

        current_bool = False
        if current_value in ['1', 'True', 'true', 'Да', 'да', '✅ Да']:
            current_bool = True

        self.bool_var = QButtonGroup(self)

        true_radio = QRadioButton("✅ Да")
        false_radio = QRadioButton("❌ Нет")

        if current_bool:
            true_radio.setChecked(True)
        else:
            false_radio.setChecked(True)

        self.bool_var.addButton(true_radio, 1)
        self.bool_var.addButton(false_radio, 0)

        radio_layout = QHBoxLayout()
        radio_layout.addWidget(true_radio)
        radio_layout.addWidget(false_radio)

        layout.addLayout(radio_layout)

        buttons_layout = QHBoxLayout()
        ok_btn = QPushButton("✅ OK")
        ok_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("❌ Отмена")
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addWidget(ok_btn)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)

    def get_value(self):
        return "True" if self.bool_var.checkedId() == 1 else "False"


class PhotoDialog(QDialog):
    def __init__(self, parent, column_name):
        super().__init__(parent)
        self.setWindowTitle(f"Добавить фото - {column_name}")
        self.setGeometry(300, 300, 500, 400)

        self.image_data = None

        layout = QVBoxLayout(self)

        label = QLabel("📸 Добавление фотографии")
        label.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addWidget(label)

        # Превью
        self.preview_label = QLabel("Выберите изображение для предпросмотра")
        self.preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.preview_label.setMinimumHeight(200)
        layout.addWidget(self.preview_label)

        # Информация о файле
        self.info_label = QLabel("")
        layout.addWidget(self.info_label)

        # Кнопки
        buttons_layout = QHBoxLayout()
        select_btn = QPushButton("📁 Выбрать файл")
        select_btn.clicked.connect(self.load_image)
        save_btn = QPushButton("✅ Сохранить фото")
        save_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("❌ Отмена")
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addWidget(select_btn)
        buttons_layout.addWidget(save_btn)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)

        # Подсказки
        tips_label = QLabel("💡 Поддерживаемые форматы: PNG, JPG, JPEG, GIF, BMP\n💡 Рекомендуемый размер: до 5 МБ")
        tips_label.setStyleSheet("color: gray; font-size: 10px;")
        layout.addWidget(tips_label)

    def load_image(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите изображение",
            "",
            "Изображения (*.png *.jpg *.jpeg *.gif *.bmp);;Все файлы (*.*)"
        )

        if file_path:
            try:
                with open(file_path, 'rb') as f:
                    self.image_data = f.read()

                # Показываем предпросмотр
                pixmap = QPixmap(file_path)
                if not pixmap.isNull():
                    scaled_pixmap = pixmap.scaled(300, 300, Qt.AspectRatioMode.KeepAspectRatio)
                    self.preview_label.setPixmap(scaled_pixmap)

                    # Информация о файле
                    file_info = f"Файл: {os.path.basename(file_path)}\nРазмер: {len(self.image_data)} байт"
                    self.info_label.setText(file_info)

            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить изображение: {e}")

    def get_image_data(self):
        return self.image_data


class ImageViewDialog(QDialog):
    def __init__(self, parent, column_name, image_data, record_info=""):
        super().__init__(parent)
        self.setWindowTitle(f"Фото - {column_name} {record_info}")
        self.setGeometry(100, 100, 800, 600)

        self.image_data = image_data

        layout = QVBoxLayout(self)

        # Изображение
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.image_label.setScaledContents(False)

        # Прокрутка для больших изображений
        scroll_area = QScrollArea()
        scroll_area.setWidget(self.image_label)
        scroll_area.setWidgetResizable(True)
        layout.addWidget(scroll_area)

        # Загружаем изображение
        self.load_image()

        # Информация
        info_label = QLabel(
            f"Размер: {self.original_width}x{self.original_height} пикселей | Объем: {len(image_data)} байт")
        layout.addWidget(info_label)

        # Кнопки
        buttons_layout = QHBoxLayout()
        save_btn = QPushButton("💾 Сохранить фото")
        save_btn.clicked.connect(self.save_image)
        print_btn = QPushButton("🖨️ Печать")
        print_btn.clicked.connect(self.print_image)
        close_btn = QPushButton("❌ Закрыть")
        close_btn.clicked.connect(self.close)

        buttons_layout.addWidget(save_btn)
        buttons_layout.addWidget(print_btn)
        buttons_layout.addStretch()
        buttons_layout.addWidget(close_btn)

        layout.addLayout(buttons_layout)

    def load_image(self):
        """Загрузка изображения"""
        try:
            image = Image.open(BytesIO(self.image_data))
            self.original_width, self.original_height = image.size

            # Конвертируем в QImage
            if image.mode == 'RGBA':
                qimage = QImage(image.tobytes(), image.width, image.height, QImage.Format.Format_RGBA8888)
            else:
                rgb_image = image.convert('RGB')
                qimage = QImage(rgb_image.tobytes(), rgb_image.width, rgb_image.height, QImage.Format.Format_RGB888)

            pixmap = QPixmap.fromImage(qimage)
            self.image_label.setPixmap(pixmap)

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить изображение: {e}")

    def save_image(self):
        """Сохранение изображения"""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить изображение",
            "",
            "PNG files (*.png);;JPEG files (*.jpg);;All files (*.*)"
        )

        if file_path:
            try:
                with open(file_path, 'wb') as f:
                    f.write(self.image_data)
                QMessageBox.information(self, "Успех", f"Изображение сохранено:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка сохранения: {e}")

    def print_image(self):
        """Печать изображения"""
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        dialog = QPrintDialog(printer, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            painter = QPainter(printer)
            pixmap = self.image_label.pixmap()
            if pixmap:
                painter.drawPixmap(0, 0, pixmap)
            painter.end()


class AddColumnDialog(QDialog):
    def __init__(self, parent, table_name):
        super().__init__(parent)
        self.setWindowTitle("Добавить колонку")
        self.setGeometry(300, 300, 400, 300)

        layout = QVBoxLayout(self)

        label = QLabel(f"Добавить колонку в таблицу '{table_name}'")
        label.setStyleSheet("font-weight: bold; font-size: 12px;")
        layout.addWidget(label)

        # Имя колонки
        layout.addWidget(QLabel("Имя колонки:"))
        self.column_name_edit = QLineEdit()
        layout.addWidget(self.column_name_edit)

        # Тип данных
        layout.addWidget(QLabel("Тип данных:"))
        self.type_combo = QComboBox()
        self.type_combo.addItems(["TEXT", "INTEGER", "REAL", "BOOLEAN", "BLOB"])
        layout.addWidget(self.type_combo)

        # Значение по умолчанию
        layout.addWidget(QLabel("Значение по умолчанию (необязательно):"))
        self.default_edit = QLineEdit()
        layout.addWidget(self.default_edit)

        # Подсказки
        help_label = QLabel(
            "💡 TEXT - текст\n💡 INTEGER - целые числа\n💡 REAL - дробные числа\n💡 BOOLEAN - да/нет\n💡 BLOB - фото и файлы")
        help_label.setStyleSheet("color: gray; font-size: 10px;")
        layout.addWidget(help_label)

        # Кнопки
        buttons_layout = QHBoxLayout()
        add_btn = QPushButton("✅ Добавить")
        add_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("❌ Отмена")
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addWidget(add_btn)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)

    def get_data(self):
        column_name = self.column_name_edit.text().strip()
        column_type = self.type_combo.currentText()
        default_value = self.default_edit.text().strip()
        return column_name, column_type, default_value if default_value else None


class MultiTableSelectDialog(QDialog):
    def __init__(self, parent, available_tables):
        super().__init__(parent)
        self.setWindowTitle("Выбор таблиц для соединения")
        self.setGeometry(300, 300, 400, 500)

        self.selected_tables = []
        self.available_tables = available_tables

        layout = QVBoxLayout(self)

        label = QLabel("🔗 Выберите таблицы для соединения")
        label.setStyleSheet("font-weight: bold; font-size: 12px;")
        layout.addWidget(label)

        # Список таблиц с чекбоксами
        self.checkboxes = []
        for table in available_tables:
            checkbox = QCheckBox(table)
            self.checkboxes.append(checkbox)
            layout.addWidget(checkbox)

        layout.addStretch()

        # Кнопки выбора всех/снятия всех
        select_buttons_layout = QHBoxLayout()
        select_all_btn = QPushButton("✅ Выбрать все")
        select_all_btn.clicked.connect(self.select_all)
        deselect_all_btn = QPushButton("❌ Снять все")
        deselect_all_btn.clicked.connect(self.deselect_all)

        select_buttons_layout.addWidget(select_all_btn)
        select_buttons_layout.addWidget(deselect_all_btn)
        layout.addLayout(select_buttons_layout)

        # Информация
        info_label = QLabel("ℹ️ Будут автоматически соединены по общим полям")
        info_label.setStyleSheet("color: gray; font-size: 10px;")
        layout.addWidget(info_label)

        # Кнопки диалога
        dialog_buttons_layout = QHBoxLayout()
        join_btn = QPushButton("🔗 Соединить выбранные")
        join_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("❌ Отмена")
        cancel_btn.clicked.connect(self.reject)

        dialog_buttons_layout.addWidget(join_btn)
        dialog_buttons_layout.addWidget(cancel_btn)
        layout.addLayout(dialog_buttons_layout)

    def select_all(self):
        for checkbox in self.checkboxes:
            checkbox.setChecked(True)

    def deselect_all(self):
        for checkbox in self.checkboxes:
            checkbox.setChecked(False)

    def get_selected_tables(self):
        selected = []
        for i, checkbox in enumerate(self.checkboxes):
            if checkbox.isChecked():
                selected.append(self.available_tables[i])
        return selected


class ExcelImportDialog(QDialog):
    def __init__(self, parent, excel_columns):
        super().__init__(parent)
        self.setWindowTitle("Импорт из Excel")
        self.setGeometry(300, 300, 500, 400)

        layout = QVBoxLayout(self)

        label = QLabel("📥 Импорт данных из Excel")
        label.setStyleSheet("font-weight: bold; font-size: 12px;")
        layout.addWidget(label)

        # Информация
        info_label = QLabel(f"Колонки в Excel: {len(excel_columns)}")
        layout.addWidget(info_label)

        # Предупреждение
        warning_label = QLabel("⚠️ Убедитесь, что структура Excel соответствует структуре таблицы!")
        warning_label.setStyleSheet("color: orange; font-size: 10px;")
        layout.addWidget(warning_label)

        # Список колонок
        layout.addWidget(QLabel("Колонки в файле Excel:"))

        list_widget = QListWidget()
        for col in excel_columns:
            list_widget.addItem(col)
        layout.addWidget(list_widget)

        # Кнопки
        buttons_layout = QHBoxLayout()
        import_btn = QPushButton("✅ Импортировать")
        import_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("❌ Отмена")
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addWidget(import_btn)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)


class JoinTablesDialog(QDialog):
    def __init__(self, parent, current_table, connection):
        super().__init__(parent)
        self.setWindowTitle("Соединить таблицы")
        self.setGeometry(300, 300, 500, 400)

        self.current_table = current_table
        self.connection = connection

        layout = QVBoxLayout(self)

        label = QLabel("🔗 Соединение таблиц")
        label.setStyleSheet("font-weight: bold; font-size: 12px;")
        layout.addWidget(label)

        layout.addWidget(QLabel(f"Основная таблица: {current_table}"))

        # Вторая таблица
        layout.addWidget(QLabel("Таблица для соединения:"))
        self.table2_combo = QComboBox()
        self.load_tables()
        layout.addWidget(self.table2_combo)

        # Атрибуты
        layout.addWidget(QLabel("Атрибут из основной таблицы:"))
        self.attr1_combo = QComboBox()
        self.load_attributes(current_table, self.attr1_combo)
        layout.addWidget(self.attr1_combo)

        layout.addWidget(QLabel("Атрибут из второй таблицы:"))
        self.attr2_combo = QComboBox()
        layout.addWidget(self.attr2_combo)

        # Тип соединения
        layout.addWidget(QLabel("Тип соединения:"))
        self.join_type_combo = QComboBox()
        self.join_type_combo.addItems(["INNER JOIN", "LEFT JOIN"])
        layout.addWidget(self.join_type_combo)

        # Предпросмотр запроса
        layout.addWidget(QLabel("Предпросмотр запроса:"))
        self.query_preview = QTextEdit()
        self.query_preview.setReadOnly(True)
        self.query_preview.setMaximumHeight(100)
        layout.addWidget(self.query_preview)

        # Подключение сигналов
        self.table2_combo.currentTextChanged.connect(self.update_second_table_attributes)
        self.attr1_combo.currentTextChanged.connect(self.update_query_preview)
        self.attr2_combo.currentTextChanged.connect(self.update_query_preview)
        self.join_type_combo.currentTextChanged.connect(self.update_query_preview)

        # Кнопки
        buttons_layout = QHBoxLayout()
        join_btn = QPushButton("🔗 Соединить")
        join_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("❌ Отмена")
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addWidget(join_btn)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)

        self.update_query_preview()

    def load_tables(self):
        """Загрузка списка таблиц"""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = cursor.fetchall()

            for table in tables:
                if table[0] != self.current_table and table[0] != "sqlite_sequence":
                    self.table2_combo.addItem(table[0])

            if self.table2_combo.count() > 0:
                self.table2_combo.setCurrentIndex(0)
                self.update_second_table_attributes()

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки таблиц: {e}")

    def load_attributes(self, table_name, combo_box):
        """Загрузка атрибутов таблицы"""
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info('{table_name}')")
            columns = cursor.fetchall()

            combo_box.clear()
            for col in columns:
                combo_box.addItem(col[1])

            if combo_box.count() > 0:
                combo_box.setCurrentIndex(0)

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки атрибутов: {e}")

    def update_second_table_attributes(self):
        """Обновление атрибутов второй таблицы"""
        table2 = self.table2_combo.currentText()
        if table2:
            self.load_attributes(table2, self.attr2_combo)
            self.update_query_preview()

    def update_query_preview(self):
        """Обновление предпросмотра запроса"""
        table2 = self.table2_combo.currentText()
        attr1 = self.attr1_combo.currentText()
        attr2 = self.attr2_combo.currentText()
        join_type = self.join_type_combo.currentText().split()[0]

        if table2 and attr1 and attr2:
            query = f"SELECT *\nFROM {self.current_table}\n{join_type} JOIN {table2}\nON {self.current_table}.{attr1} = {table2}.{attr2}"
            self.query_preview.setText(query)

    def get_data(self):
        table2 = self.table2_combo.currentText()
        attr1 = self.attr1_combo.currentText()
        attr2 = self.attr2_combo.currentText()
        join_type = self.join_type_combo.currentText().split()[0]
        return table2, attr1, attr2, join_type


class SelectAttributesDialog(QDialog):
    def __init__(self, parent, all_columns, selected_attributes):
        super().__init__(parent)
        self.setWindowTitle("Выбор атрибутов для отображения")
        self.setGeometry(300, 300, 500, 600)

        self.all_columns = all_columns
        self.selected_attributes = selected_attributes.copy()

        layout = QVBoxLayout(self)

        label = QLabel("👁️ Выберите атрибуты для отображения")
        label.setStyleSheet("font-weight: bold; font-size: 12px;")
        layout.addWidget(label)

        # Список с чекбоксами
        self.checkboxes = {}

        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        row = 0
        for table_name, columns in all_columns.items():
            table_label = QLabel(f"📋 Таблица: {table_name}")
            table_label.setStyleSheet("font-weight: bold;")
            scroll_layout.addWidget(table_label)

            for column in columns:
                full_attr_name = f"{table_name}.{column}"
                checkbox = QCheckBox(column)
                checkbox.setChecked(full_attr_name in selected_attributes)
                self.checkboxes[full_attr_name] = checkbox
                scroll_layout.addWidget(checkbox)

            scroll_layout.addSpacing(10)

        scroll_area.setWidget(scroll_widget)
        layout.addWidget(scroll_area)

        # Кнопки управления
        manage_buttons_layout = QHBoxLayout()
        select_all_btn = QPushButton("✅ Выбрать все")
        select_all_btn.clicked.connect(self.select_all)
        deselect_all_btn = QPushButton("❌ Снять все")
        deselect_all_btn.clicked.connect(self.deselect_all)

        manage_buttons_layout.addWidget(select_all_btn)
        manage_buttons_layout.addWidget(deselect_all_btn)
        layout.addLayout(manage_buttons_layout)

        # Кнопки диалога
        dialog_buttons_layout = QHBoxLayout()
        apply_btn = QPushButton("✅ Применить")
        apply_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("❌ Отмена")
        cancel_btn.clicked.connect(self.reject)
        show_all_btn = QPushButton("👁️ Показать все")
        show_all_btn.clicked.connect(self.show_all)

        dialog_buttons_layout.addWidget(apply_btn)
        dialog_buttons_layout.addWidget(cancel_btn)
        dialog_buttons_layout.addWidget(show_all_btn)
        layout.addLayout(dialog_buttons_layout)

    def select_all(self):
        for checkbox in self.checkboxes.values():
            checkbox.setChecked(True)

    def deselect_all(self):
        for checkbox in self.checkboxes.values():
            checkbox.setChecked(False)

    def show_all(self):
        self.selected_attributes = []
        self.accept()

    def get_selected_attributes(self):
        selected = []
        for attr_name, checkbox in self.checkboxes.items():
            if checkbox.isChecked():
                selected.append(attr_name)
        return selected


class CreateTableDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Создать таблицу")
        self.setGeometry(300, 300, 600, 500)

        self.columns = []

        layout = QVBoxLayout(self)

        label = QLabel("📊 Создание новой таблицы")
        label.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addWidget(label)

        # Название таблицы
        layout.addWidget(QLabel("Название таблицы:"))
        self.table_name_edit = QLineEdit()
        layout.addWidget(self.table_name_edit)

        # Колонки
        columns_group = QGroupBox("📋 Колонки таблицы")
        columns_layout = QVBoxLayout()

        self.columns_list = QListWidget()
        columns_layout.addWidget(self.columns_list)

        # Кнопки управления колонками
        column_buttons_layout = QHBoxLayout()
        add_column_btn = QPushButton("➕ Добавить колонку")
        add_column_btn.clicked.connect(self.add_column_dialog)
        remove_column_btn = QPushButton("🗑️ Удалить колонку")
        remove_column_btn.clicked.connect(self.remove_column)

        column_buttons_layout.addWidget(add_column_btn)
        column_buttons_layout.addWidget(remove_column_btn)
        columns_layout.addLayout(column_buttons_layout)

        columns_group.setLayout(columns_layout)
        layout.addWidget(columns_group)

        # Кнопки диалога
        dialog_buttons_layout = QHBoxLayout()
        create_btn = QPushButton("✅ Создать таблицу")
        create_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("❌ Отмена")
        cancel_btn.clicked.connect(self.reject)

        dialog_buttons_layout.addWidget(create_btn)
        dialog_buttons_layout.addWidget(cancel_btn)
        layout.addLayout(dialog_buttons_layout)

    def add_column_dialog(self):
        dialog = AddColumnDialog(self, "")
        if dialog.exec():
            column_name, column_type, default_value = dialog.get_data()
            if column_name:
                column = {"name": column_name, "type": column_type}
                self.columns.append(column)
                display_text = f"{column_name} ({column_type})"
                if default_value:
                    display_text += f" [по умолчанию: {default_value}]"
                self.columns_list.addItem(display_text)

    def remove_column(self):
        current_row = self.columns_list.currentRow()
        if current_row >= 0:
            self.columns_list.takeItem(current_row)
            self.columns.pop(current_row)

    def get_data(self):
        table_name = self.table_name_edit.text().strip()
        return table_name, self.columns


class AddRecordDialog(QDialog):
    def __init__(self, parent, table_name, connection):
        super().__init__(parent)
        self.setWindowTitle("Добавить запись")
        self.setGeometry(300, 300, 400, 500)

        self.table_name = table_name
        self.connection = connection
        self.entries = {}

        layout = QVBoxLayout(self)

        label = QLabel(f"➕ Добавить запись в '{table_name}'")
        label.setStyleSheet("font-weight: bold; font-size: 12px;")
        layout.addWidget(label)

        # Прокручиваемая область для полей
        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info('{table_name}')")
            columns = cursor.fetchall()

            for i, column in enumerate(columns):
                col_name = column[1]
                col_type = column[2]

                row_layout = QHBoxLayout()
                row_layout.addWidget(QLabel(f"{col_name} ({col_type}):"))

                if col_type.upper() == 'BOOLEAN':
                    entry = QComboBox()
                    entry.addItems(["False", "True", "0", "1", "Нет", "Да"])
                    entry.setCurrentText("False")
                else:
                    entry = QLineEdit()

                self.entries[col_name] = (entry, col_type)
                row_layout.addWidget(entry)
                scroll_layout.addLayout(row_layout)

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка получения структуры таблицы: {e}")
            self.reject()

        scroll_area.setWidget(scroll_widget)
        layout.addWidget(scroll_area)

        # Подсказка
        help_label = QLabel("Для BOOLEAN: True/1/Да или False/0/Нет")
        help_label.setStyleSheet("color: gray; font-size: 10px;")
        layout.addWidget(help_label)

        # Кнопки
        buttons_layout = QHBoxLayout()
        add_btn = QPushButton("✅ Добавить")
        add_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("❌ Отмена")
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addWidget(add_btn)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)

    def get_values(self):
        values = []
        for col_name, (entry, col_type) in self.entries.items():
            if isinstance(entry, QLineEdit):
                value = entry.text().strip()
            else:  # QComboBox
                value = entry.currentText().strip()

            if value == "":
                values.append(None)
            else:
                values.append(value)

        return values


class ExportSettingsDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Настройки экспорта")
        self.setGeometry(300, 300, 400, 300)

        layout = QVBoxLayout(self)

        label = QLabel("⚙️ Настройки экспорта фото")
        label.setStyleSheet("font-weight: bold; font-size: 12px;")
        layout.addWidget(label)

        # Опции
        self.include_images_check = QCheckBox("Включать фото в Excel")
        self.include_images_check.setChecked(True)
        layout.addWidget(self.include_images_check)

        self.save_as_files_check = QCheckBox("Сохранять фото как отдельные файлы")
        layout.addWidget(self.save_as_files_check)

        layout.addWidget(QLabel("Размер миниатюр (пикселей):"))

        self.size_group = QButtonGroup(self)
        small_radio = QRadioButton("Маленькие (80px)")
        medium_radio = QRadioButton("Средние (100px)")
        large_radio = QRadioButton("Большие (150px)")

        self.size_group.addButton(small_radio, 80)
        self.size_group.addButton(medium_radio, 100)
        self.size_group.addButton(large_radio, 150)

        medium_radio.setChecked(True)

        size_layout = QHBoxLayout()
        size_layout.addWidget(small_radio)
        size_layout.addWidget(medium_radio)
        size_layout.addWidget(large_radio)
        layout.addLayout(size_layout)

        # Кнопки
        buttons_layout = QHBoxLayout()
        proceed_btn = QPushButton("✅ Продолжить")
        proceed_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("❌ Отмена")
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addWidget(proceed_btn)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)

    def get_settings(self):
        return {
            'include_images': self.include_images_check.isChecked(),
            'save_as_files': self.save_as_files_check.isChecked(),
            'image_size': self.size_group.checkedId()
        }


def main():
    app = QApplication(sys.argv)

    # Установка стиля
    app.setStyle('Fusion')

    window = ModernDatabaseApp()
    window.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()