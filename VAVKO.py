ИЗМЕНЕНИЕ
import sys
import os
import sqlite3
import tempfile
import shutil
from datetime import datetime
from io import BytesIO

import pandas as pd
from PyQt6.QtWidgets import *
from PyQt6.QtCore import *
from PyQt6.QtGui import *
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
from PyQt6.QtGui import QIcon 

from PIL import Image
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Константы
CELL_WIDTH = 120
CELL_HEIGHT = 100
PHOTO_COLUMN_WIDTH = 120
TEXT_COLUMN_WIDTH = 100
ZOOM_MIN = 10
ZOOM_MAX = 300
ZOOM_DEFAULT = 100

# Единая стилизация: строгий, читаемый интерфейс
APP_STYLESHEET = """
    QMainWindow, QWidget { background-color: #f5f6f8; }
    
    QLabel { color: #1a1d21; font-size: 9pt; font-family: 'Segoe UI', Arial, sans-serif; }
    
    QGroupBox {
        font-size: 9pt; font-weight: bold; color: #2c3e50;
        border: 1px solid #cbd5e0; border-radius: 4px;
        margin-top: 10px; padding-top: 10px; padding: 8px;
        background-color: #ffffff;
    }
    QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 6px; }
    
    QPushButton {
        font-size: 7pt; font-family: 'Segoe UI', Arial, sans-serif;
        min-height: 18px; padding: 2px 6px; border-radius: 3px;
        border: 1px solid transparent;
    }
    QPushButton:focus { outline: none; border: 1px solid #2c5282; }
    
    QLineEdit, QComboBox, QTextEdit, QSpinBox {
        font-size: 9pt; padding: 4px 6px; border-radius: 3px;
        border: 1px solid #cbd5e0; background: white;
        font-family: 'Segoe UI', Arial, sans-serif; color: #1a1d21;
    }
    QLineEdit:focus, QComboBox:focus, QTextEdit:focus { border-color: #2c5282; }
    
    QListWidget {
        font-size: 9pt; padding: 2px; border-radius: 3px;
        border: 1px solid #cbd5e0; background: white;
        font-family: 'Segoe UI', Arial, sans-serif;
    }
    QListWidget::item { padding: 4px; min-height: 18px; }
    QListWidget::item:selected { background-color: #2c5282; color: white; }
    QListWidget::item:hover:!selected { background-color: #edf2f7; }
    
    QTableWidget {
        font-size: 9pt; gridline-color: #e2e8f0;
        background: white; color: #1a1d21;
        font-family: 'Segoe UI', Arial, sans-serif;
        alternate-background-color: #f7fafc;
    }
    QTableWidget::item { padding: 4px; }
    QTableWidget::item:selected { background-color: #2c5282; color: white; }
    QHeaderView::section {
        font-size: 9pt; font-weight: bold; color: #2c3e50;
        background-color: #edf2f7; padding: 4px; border: none; border-bottom: 1px solid #cbd5e0;
    }
    
    QStatusBar {
        font-size: 9pt; color: #4a5568; background-color: #edf2f7;
        border-top: 1px solid #cbd5e0; padding: 2px;
    }
    
    QPushButton { min-width: 50px; }
    QComboBox { min-width: 100px; }
"""
def applyTextFit(widget):
    """Включает перенос текста и подгонку, чтобы не обрезался"""
    for lb in widget.findChildren(QLabel):
        lb.setWordWrap(True)
    for btn in widget.findChildren(QPushButton):
        # QPushButton doesn't support setWordWrap; avoid crashing.
        if hasattr(btn, "setWordWrap"):
            btn.setWordWrap(True)
    for cb in widget.findChildren(QComboBox):
        cb.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents)
        cb.setMinimumContentsLength(12)
    for chk in widget.findChildren(QCheckBox):
        # QCheckBox also doesn't support setWordWrap in Qt; keep safe.
        if hasattr(chk, "setWordWrap"):
            chk.setWordWrap(True)

class ImageWidget(QWidget):
    """Виджет для отображения миниатюры изображения"""
    clicked = pyqtSignal(int, int)
    rightClicked = pyqtSignal(int, int)
    
    def __init__(self, image_data, row, col):
        super().__init__()
        self.row = row
        self.col = col
        self.image_data = image_data
        self.pixmap = None
        self._qimage_buffer = None
        self.initUI()
        
    def initUI(self):
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.setStyleSheet("background-color: #ffffff; border: 1px solid #cbd5e0; border-radius: 4px;")
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.label = QLabel()
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.label.setCursor(Qt.CursorShape.PointingHandCursor)
        self.label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        
        self.loadImage()
        layout.addWidget(self.label)
        
    def loadImage(self):
        try:
            img = Image.open(BytesIO(self.image_data))
            if img.mode == 'RGBA':
                self._qimage_buffer = img.tobytes()
                qimg = QImage(self._qimage_buffer, img.width, img.height,
                              img.width * 4, QImage.Format.Format_RGBA8888)
            else:
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                self._qimage_buffer = img.tobytes()
                qimg = QImage(self._qimage_buffer, img.width, img.height,
                              img.width * 3, QImage.Format.Format_RGB888)
            
            self.pixmap = QPixmap.fromImage(qimg)
            self.updateScale()
        except:
            self.label.setText("⚠️ Ошибка")
            self.label.setStyleSheet("color: red; font-size: 10px;")
    
    def updateScale(self):
        if self.pixmap and not self.pixmap.isNull():
            w = self.label.width() - 4
            h = self.label.height() - 4
            if w > 0 and h > 0:
                dpr = self.label.devicePixelRatioF()
                target_w = max(1, int(w * dpr))
                target_h = max(1, int(h * dpr))
                scaled = self.pixmap.scaled(target_w, target_h, Qt.AspectRatioMode.KeepAspectRatio,
                                            Qt.TransformationMode.SmoothTransformation)
                scaled.setDevicePixelRatio(dpr)
                self.label.setPixmap(scaled)
    
    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.updateScale()
    
    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit(self.row, self.col)
        elif event.button() == Qt.MouseButton.RightButton:
            self.rightClicked.emit(self.row, self.col)
        super().mousePressEvent(event)
    
    def enterEvent(self, event):
        self.setStyleSheet("background-color: #edf2f7; border: 2px solid #2c5282; border-radius: 4px;")
        
    def leaveEvent(self, event):
        self.setStyleSheet("background-color: #ffffff; border: 1px solid #cbd5e0; border-radius: 4px;")


class ModernDatabaseApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.column_mapping = {}
        self.image_columns = []
        self.db_name = None
        self.current_table = None
        self.connection = None
        self.joined_tables = []
        self.selected_attributes = []
        self.table_joins = {}
        self.russian_font_registered = False
        self.initUI()
        self.selectDatabase()
        
    def initUI(self):
        self.setWindowTitle("Database Manager")
        self.setGeometry(100, 100, 1200, 700)
        self.setMinimumSize(800, 500)
        #self.setWindowIcon(QIcon('icon.png')) 
        
        
        # Регистрируем русский шрифт для PDF
        self.registerRussianFont()
        
        # Шрифт: Segoe UI — читаемый для всех возрастов
        app_font = QFont("Segoe UI", 9)
        app_font.setFamily("Segoe UI")
        app_font.setPointSize(9)
        self.setFont(app_font)
        
        # Центральный виджет
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Заголовок
        header = self.createHeader()
        main_layout.addWidget(header)
        
        # Панель действий
        actions = self.createActions()
        main_layout.addWidget(actions)
        
        # Основной контент
        content = self.createContent()
        main_layout.addWidget(content)
        
        # Статус бар
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.status.showMessage("Готов к работе")
        
        applyTextFit(self)
        self.setupHotkeys()

    def printToPrinter(self):
        """Печать на физический принтер"""
        if not self.current_table and not self.joined_tables:
            QMessageBox.warning(self, "Предупреждение", "Нет данных")
            return
        
        try:
            query, cols = self.buildQuery()
            cursor = self.connection.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            
            if not rows:
                QMessageBox.information(self, "Информация", "Нет данных")
                return
            
            # ---- УНИВЕРСАЛЬНЫЙ БЛОК (работает во всех версиях PyQt6) ----
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            
            # Задаём страницу через QPageSize (существует в вашей версии)
            printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
            # Задаём ориентацию через QPageLayout (существует в вашей версии)
            printer.setPageOrientation(QPageLayout.Orientation.Landscape)
            # -------------------------------------------------------------
            
            # Открываем диалог печати
            dialog = QPrintDialog(printer, self)
            dialog.setWindowTitle("Печать отчета")
            
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return
            
            # Создаём Painter для рисования на принтере
            painter = QPainter(printer)
            painter.begin(printer)
            
            # Получаем размеры страницы принтера в пикселях
            page_rect = printer.pageRect(QPrinter.Unit.DevicePixel)
            page_width = page_rect.width()
            page_height = page_rect.height()
            
            margin = int(page_width * 0.05)  # 5% от ширины страницы
            table_width = page_width - 2 * margin
            
            # Рассчитываем ширину колонок (в пикселях принтера)
            num_cols = len(cols)
            col_width = table_width / num_cols
            if col_width < 150: col_width = 150  # минимальная ширина
            
            # Координаты начала рисования
            y = margin + 30
            
            # --- ЗАГОЛОВОК ОТЧЕТА ---
            painter.setPen(QColor(0, 0, 0))
            painter.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
            painter.drawText(margin, y, f"Отчет: {self.current_table}")
            y += 40
            
            painter.setFont(QFont("Segoe UI", 10))
            painter.drawText(margin, y, f"База данных: {os.path.basename(self.db_name)}")
            y += 22
            painter.drawText(margin, y, f"Дата создания: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            y += 40
            
            # --- ПОДГОТОВКА ДАННЫХ ---
            formatted_rows = []
            img_cols = [c for c in cols if self.isImageColumn(c)]
            
            for r, row in enumerate(rows):
                formatted_row = []
                max_height = 40  # Минимальная высота строки в пикселях принтера
                
                for c, val in enumerate(row):
                    name = cols[c]
                    cell_data = {'text_lines': [], 'image_data': None, 'height': 40}
                    
                    if name in img_cols and val and isinstance(val, bytes) and self.isValidImage(val):
                        cell_data['image_data'] = val
                        cell_data['height'] = 150  # Фото требует высоты
                    
                    elif val is not None and not isinstance(val, bool):
                        text = str(val)
                        # Перенос текста (приблизительный расчет)
                        char_width = 9  # средняя ширина символа в пикселях принтера
                        max_chars = int(col_width / char_width) - 2
                        lines = self.wrap_text(text, max_chars)
                        cell_data['text_lines'] = lines
                        h = len(lines) * 15 + 8
                        if h > cell_data['height']:
                            cell_data['height'] = h
                    
                    else:
                        if isinstance(val, bool):
                            cell_data['text_lines'] = ["✅ Да" if val else "❌ Нет"]
                        else:
                            cell_data['text_lines'] = [""]
                        cell_data['height'] = 35
                    
                    if cell_data['height'] > max_height:
                        max_height = cell_data['height']
                    
                    formatted_row.append(cell_data)
                
                for cell in formatted_row:
                    cell['height'] = max_height
                
                formatted_rows.append({'cells': formatted_row, 'height': max_height})
            
            # --- ОТРИСОВКА ТАБЛИЦЫ НА ПРИНТЕРЕ ---
            header_height = 35
            
            # 1. Рисуем заголовки колонок
            painter.setPen(Qt.PenStyle.SolidLine)
            x = margin
            
            painter.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
            
            # Верхняя линия
            painter.drawLine(margin, y, margin + table_width, y)
            y -= header_height
            painter.drawLine(margin, y, margin + table_width, y)
            
            for i, name in enumerate(cols):
                painter.drawLine(x, y + header_height, x, y)
                # Отрисовка текста с отступом
                painter.drawText(x + 6, y + 12, str(name))
                x += col_width
            painter.drawLine(margin + table_width, y + header_height, margin + table_width, y)
            
            # 2. Рисуем строки данных
            painter.setFont(QFont("Segoe UI", 8))
            
            for row_data in formatted_rows:
                row_height = row_data['height'] + 8
                
                # Если не хватает места, создаем новую страницу
                if y - row_height < margin:
                    printer.newPage()
                    y = page_height - margin - 20
                    x = margin
                    painter.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
                    painter.drawLine(margin, y, margin + table_width, y)
                    y -= header_height
                    painter.drawLine(margin, y, margin + table_width, y)
                    for i, name in enumerate(cols):
                        painter.drawLine(x, y + header_height, x, y)
                        painter.drawText(x + 6, y + 12, str(name))
                        x += col_width
                    painter.drawLine(margin + table_width, y + header_height, margin + table_width, y)
                    painter.setFont(QFont("Segoe UI", 8))
                    y -= 6
                
                y -= row_height
                painter.drawLine(margin, y, margin + table_width, y)
                
                x = margin
                for i, cell in enumerate(row_data['cells']):
                    painter.drawLine(x, y + row_height, x, y)
                    
                    if cell['image_data']:
                        # ЗАГРУЗКА И РИСОВАНИЕ ФОТО НА ПРИНТЕР
                        try:
                            qimg = QImage()
                            qimg.loadFromData(cell['image_data'])
                            pix = QPixmap.fromImage(qimg)
                            
                            if not pix.isNull():
                                max_w = int(col_width - 12)
                                max_h = int(row_height - 12)
                                
                                # Масштабируем для принтера
                                scaled = pix.scaled(max_w, max_h, 
                                                    Qt.AspectRatioMode.KeepAspectRatio,
                                                    Qt.TransformationMode.SmoothTransformation)
                                
                                img_x = int(x + (col_width - scaled.width()) / 2)
                                img_y = int(y + (row_height - scaled.height()) / 2)
                                painter.drawPixmap(img_x, img_y, scaled)
                            else:
                                painter.drawText(x + 6, y + int(row_height/2), "⚠️ Ошибка фото")
                        except Exception as e:
                            painter.drawText(x + 6, y + int(row_height/2), "⚠️ Ошибка")
                    
                    elif cell['text_lines']:
                        lines = cell['text_lines']
                        line_height = 14
                        total_text_height = len(lines) * line_height
                        start_y = int(y + (row_height - total_text_height) / 2 + (line_height - 2))
                        
                        for j, line in enumerate(lines):
                            painter.drawText(x + 6, start_y + j * line_height, line)
                    
                    x += col_width
                
                # Правая граница строки
                painter.drawLine(margin + table_width, y + row_height, margin + table_width, y)
            
            painter.end()  # Завершаем рисование на принтере
            self.updateStatus("✅ Отправлено на печать")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка печати", str(e))
            print(f"Ошибка при печати: {e}")
    def fixPhotoRowHeights(self):
        """Принудительно выставляет высоту строк с фото, чтобы виджеты не срезались"""
        if not hasattr(self, 'table') or not self.image_columns:
            return
            
        # Получаем индекс колонки с фото (берём первую найденную)
        photo_col_index = -1
        for i in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(i)
            if header and header.text() in self.image_columns:
                photo_col_index = i
                break
                
        if photo_col_index == -1:
            return

        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, photo_col_index)
            if w and isinstance(w, ImageWidget):
                # Если в ячейке есть виджет с фото, запрещаем таблице сжимать эту строку
                # Размер берем из константы (100px) и добавляем запас на рамки
                self.table.setRowHeight(r, CELL_HEIGHT + 4)        
    def registerRussianFont(self):
        """Регистрация русского шрифта для PDF"""
        # Пути к шрифтам с кириллицей
        font_paths = [
            "C:/Windows/Fonts/arial.ttf",           # Windows Arial
            "C:/Windows/Fonts/times.ttf",           # Windows Times New Roman
            "C:/Windows/Fonts/segoeui.ttf",         # Windows Segoe UI
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",  # Linux
            "/System/Library/Fonts/Arial.ttf",       # macOS
        ]
        
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('RussianFont', font_path))
                    self.russian_font_registered = True
                    print(f"Зарегистрирован шрифт: {font_path}")
                    break
                except Exception as e:
                    print(f"Ошибка регистрации шрифта {font_path}: {e}")
                    continue
        
        if not self.russian_font_registered:
            print("Предупреждение: Не найден шрифт с поддержкой кириллицы")
    
    def wrap_text(self, text, max_chars):
        """Разбивает текст на строки по максимальному количеству символов"""
        if len(text) <= max_chars:
            return [text]
        
        lines = []
        words = text.split()
        current_line = ""
        
        for word in words:
            if len(current_line) + len(word) + 1 <= max_chars:
                if current_line:
                    current_line += " " + word
                else:
                    current_line = word
            else:
                if current_line:
                    lines.append(current_line)
                current_line = word
        
        if current_line:
            lines.append(current_line)
        
        return lines
    
    def createHeader(self):
        widget = QWidget()
        widget.setStyleSheet("background: white; border-radius: 8px; padding: 12px; border: 1px solid #cbd5e0;")
        layout = QHBoxLayout(widget)
        
        title = QLabel("Database Manager")
        title_font = QFont("Segoe UI", 14, QFont.Weight.Bold)
        title.setFont(title_font)
        title.setStyleSheet("color: #2c3e50;")
        
        self.db_label = QLabel("База: не выбрана")
        self.db_label.setFont(QFont("Segoe UI", 9))
        self.db_label.setStyleSheet("color: #4a5568;")
        
        layout.addWidget(title)
        #layout.addWidget(hotkeys)
        layout.addStretch()
        layout.addWidget(self.db_label)
        
        return widget
    
    def createActions(self):
        group = QGroupBox("Быстрые действия")
        layout = QGridLayout()
        layout.setContentsMargins(4, 4, 4, 4)
        actions = [            ("📊 Создать таблицу", self.createTable, "primary"),
            ("🗑️ Удалить таблицу", self.deleteTable, "danger"),
            ("🔗 Соединение", self.quickJoin, "primary"),
            ("👁️ Атрибуты", self.selectAttributes, "secondary"),
            ("📥 Импорт Excel", self.importExcel, "success"),
            ("📤 Экспорт Excel", self.exportExcelWithPhotos, "success"),
            ("🔍 Исследовать", self.inspectDB, "primary"),
            ("🖨️ Сохранить PDF", self.printData, "warning"),
            ("🖨️ Напечатать", self.printToPrinter, "warning"),  # <--- НОВАЯ КНОПКА
            ("💾 Сменить БД", self.changeDB, "secondary"),
        ]
        
        for i, (text, callback, style) in enumerate(actions):
            btn = QPushButton(text)
            btn.setMaximumHeight(40)
            btn.clicked.connect(callback)
            self.styleButton(btn, style)
            
            btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)
            btn.setMinimumHeight(20)
            
            row = i // 5  # Теперь 5 колонок в ряду
            col = i % 5
            layout.addWidget(btn, row, col)
        
        for i in range(5):
            layout.setColumnStretch(i, 1)
            
        layout.setRowStretch(0, 1)
        layout.setRowStretch(1, 1)
            
        group.setLayout(layout)
        return group
    
    def styleButton(self, btn, style):
        colors = {
            "primary": "#2c5282",
            "secondary": "#4a5568",
            "success": "#276749",
            "danger": "#c53030",
            "warning": "#d69e2e"
        }
        hover = {
            "primary": "#234772",
            "secondary": "#3d4852",
            "success": "#22543d",
            "danger": "#9b2c2c",
            "warning": "#b7791f"
        }
        text_color = "#1a1d21" if style == "warning" else "white"
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {colors[style]};
                color: {text_color};
                border: 1px solid {hover[style]};
                padding: 6px 12px;
                border-radius: 6px;
                font-size: 9pt;
                font-family: 'Segoe UI', Arial, sans-serif;
            }}
            QPushButton:hover {{ background-color: {hover[style]}; }}
            QPushButton:focus {{ border: 2px solid #2c5282; }}
        """)
    
    def createContent(self):
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Левая панель
        left = QWidget()
        left.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(3, 3, 3, 3)
        
        # Таблицы
        tables_group = QGroupBox("Таблицы")
        tables_layout = QVBoxLayout()
        tables_layout.setContentsMargins(4, 4, 4, 4)
        
        search = QHBoxLayout()
        search.addWidget(QLabel("Поиск:"))
        self.table_search = QLineEdit()
        self.table_search.textChanged.connect(self.filterTables)
        search.addWidget(self.table_search)
        
        self.table_list = QListWidget()
        self.table_list.itemSelectionChanged.connect(self.onTableSelect)
        
        tables_layout.addLayout(search)
        tables_layout.addWidget(self.table_list)
        tables_group.setLayout(tables_layout)
        
        # Соединения
        joins_group = QGroupBox("Соединения")
        joins_layout = QVBoxLayout()
        joins_layout.setContentsMargins(4, 4, 4, 4)
        
        self.join_info = QTextEdit()
        self.join_info.setReadOnly(True)
        self.join_info.setMinimumHeight(60)
        
        btn_layout = QHBoxLayout()
        clear_btn = QPushButton("🗑️ Очистить")
        clear_btn.clicked.connect(self.clearJoins)
        clear_btn.setMinimumHeight(18)
        
        remove_btn = QPushButton("✂️ Удалить")
        remove_btn.clicked.connect(self.removeJoin)
        remove_btn.setMinimumHeight(18)
        
        advanced_btn = QPushButton("⚙️ Расширенное")
        advanced_btn.clicked.connect(self.joinTablesAdvanced)
        advanced_btn.setMinimumHeight(18)
        
        for btn in (clear_btn, remove_btn, advanced_btn):
            self.styleButton(btn, "danger" if btn == clear_btn else "secondary" if btn == remove_btn else "primary")
            btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)
            btn_layout.addWidget(btn)
            
        joins_layout.addWidget(self.join_info)
        joins_layout.addLayout(btn_layout)
        joins_group.setLayout(joins_layout)
        
        left_layout.addWidget(tables_group)
        left_layout.addWidget(joins_group)
        
        # Правая панель (Данные)
        right = QWidget()
        right.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(3, 3, 3, 3)
        
        # Инструменты
        tools_group = QGroupBox("Данные")
        tools_layout = QVBoxLayout()
        tools_layout.setContentsMargins(4, 4, 4, 4)
        
        # Сортировка
        sort_layout = QGridLayout()
        sort_layout.addWidget(QLabel("Сортировка:"), 0, 0)
        self.sort_col = QComboBox()
        self.sort_col.setMinimumWidth(80)
        sort_layout.addWidget(self.sort_col, 0, 1)
        
        self.sort_order = QComboBox()
        self.sort_order.addItems(["По возрастанию", "По убыванию"])
        self.sort_order.setMinimumWidth(80)
        sort_layout.addWidget(self.sort_order, 0, 2)
        
        apply_btn = QPushButton("🔄 Применить")
        apply_btn.clicked.connect(self.applySorting)
        apply_btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)
        apply_btn.setMinimumHeight(18)
        sort_layout.addWidget(apply_btn, 0, 3)
        
        sort_layout.setColumnStretch(0, 0)
        sort_layout.setColumnStretch(1, 1)
        sort_layout.setColumnStretch(2, 1)
        sort_layout.setColumnStretch(3, 1)
        
        tools_layout.addLayout(sort_layout)
        
        # Атрибуты
        self.attr_label = QLabel("Атрибуты: все")
        tools_layout.addWidget(self.attr_label)
        
        # Кнопки
        edit_btns = QGridLayout()
        
        add_btn_ = QPushButton("➕ Добавить запись")
        add_btn_.clicked.connect(self.addRecord)
        
        add_col = QPushButton("📝 Добавить колонку")
        add_col.clicked.connect(self.addColumn)
        
        delete_btn = QPushButton("🗑️ Удалить запись")
        delete_btn.clicked.connect(self.deleteRecord)
        
        rename_btn = QPushButton("📝 Переименовать")
        rename_btn.clicked.connect(self.renameAttribute)
        
        buttons_data = [
            (add_btn_, "success"),
            (add_col, "success"),
            (delete_btn, "danger"),
            (rename_btn, "primary")
        ]
        
        for i, (btn, style) in enumerate(buttons_data):
            self.styleButton(btn, style)
            btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)
            btn.setMinimumHeight(18)
            
            row = i // 4
            col = i % 4
            edit_btns.addWidget(btn, row, col)
        
        for i in range(4):
            edit_btns.setColumnStretch(i, 1)
            
        tools_layout.addLayout(edit_btns)
        tools_group.setLayout(tools_layout)
        
        # Таблица
        self.table = QTableWidget()
        self.table.setWordWrap(True)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.showContextMenu)
        self.table.doubleClicked.connect(self.onCellDoubleClick)
        # self.table.horizontalHeader().sectionResized.connect(self.table.resizeRowsToContents)

        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        right_layout.addWidget(tools_group)
        right_layout.addWidget(self.table)
        
        right.setMinimumWidth(200)
        
        splitter.addWidget(left)
        splitter.addWidget(right)
        splitter.setSizes([220, 800])
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        
        layout.addWidget(splitter)
        
        return widget
    
    def setupHotkeys(self):
        shortcuts = [
            ("F5", self.refreshData),
            ("Ctrl+S", self.quickSave),
            ("Delete", self.quickDelete),
            ("Ctrl+P", self.printData),
            ("Return", self.onEnter),
            ("Ctrl+I", self.quickAddPhoto),
            ("Ctrl+V", self.viewPhoto)
        ]
        for key, callback in shortcuts:
            QShortcut(QKeySequence(key), self).activated.connect(callback)
    
    def onEnter(self):
        w = self.focusWidget()
        if isinstance(w, (QLineEdit, QComboBox)):
            self.refreshData()
        elif isinstance(w, QTableWidget):
            self.editCell()
    
    def quickSave(self):
        if self.connection:
            try:
                self.connection.commit()
                self.updateStatus("💾 Сохранено!")
            except sqlite3.Error as e:
                self.updateStatus(f"❌ {e}")
    
    def quickDelete(self):
        if self.table.selectionModel().hasSelection():
            self.deleteRecord()
    
    def selectDatabase(self):
        path, _ = QFileDialog.getSaveFileName(self, "Выберите БД", "", "SQLite (*.db)")
        self.db_name = path if path else "my_database.db"
        if not self.db_name.endswith('.db'):
            self.db_name += '.db'
        self.connectToDB()
    
    def connectToDB(self):
        try:
            self.connection = sqlite3.connect(self.db_name)
            self.connection.execute("PRAGMA foreign_keys = ON")
            self.updateTableList()
            self.db_label.setText(f"База: {os.path.basename(self.db_name)}")
            self.updateStatus(f"✅ Подключено к {os.path.basename(self.db_name)}")
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def changeDB(self):
        if QMessageBox.question(self, "Смена БД", "Сменить базу?") == QMessageBox.StandardButton.Yes:
            if self.connection:
                self.connection.close()
            self.selectDatabase()
    
    def updateTableList(self):
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [t[0] for t in cursor.fetchall() if t[0] != "sqlite_sequence"]
            self.table_list.clear()
            self.table_list.addItems(tables)
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def filterTables(self):
        text = self.table_search.text().lower()
        for i in range(self.table_list.count()):
            item = self.table_list.item(i)
            item.setHidden(text not in item.text().lower())
    
    def onTableSelect(self):
        items = self.table_list.selectedItems()
        if not items:
            return
        new_table = items[0].text()
        
        if self.current_table:
            self.table_joins[self.current_table] = self.joined_tables.copy()
        
        self.current_table = new_table
        self.joined_tables = self.table_joins.get(self.current_table, [])
        self.selected_attributes.clear()
        self.updateJoinInfo()
        self.updateAttributesLabel()
        self.displayTableData()
        self.updateStatus(f"📊 {new_table}")
    
    def updateJoinInfo(self):
        if self.joined_tables:
            text = f"Основная: {self.current_table}\n"
            for i, j in enumerate(self.joined_tables, 1):
                text += f"{i}. {j['table2']}\n   {j['condition']} [{j['join_type']}]\n"
        else:
            text = "Нет соединений"
        self.join_info.setText(text)
    
    def updateAttributesLabel(self):
        if self.selected_attributes:
            attrs = ", ".join([a.split('.')[-1] for a in self.selected_attributes[:3]])
            if len(self.selected_attributes) > 3:
                attrs += f"... (+{len(self.selected_attributes)-3})"
            self.attr_label.setText(f"Атрибуты: {attrs}")
        else:
            self.attr_label.setText("Атрибуты: все")
    
    def clearJoins(self):
        self.joined_tables.clear()
        if self.current_table:
            self.table_joins[self.current_table] = []
        self.updateJoinInfo()
        if self.current_table:
            self.displayTableData()
        self.updateStatus("✅ Соединения очищены")
    
    def removeJoin(self):
        if self.joined_tables:
            removed = self.joined_tables.pop()
            self.table_joins[self.current_table] = self.joined_tables.copy()
            self.updateJoinInfo()
            self.displayTableData()
            self.updateStatus(f"✅ Удалено {removed['table2']}")
    
    def escape(self, name):
        return f'"{name}"'
    
    def buildQuery(self, sort_col=None, sort_order="ASC"):
        if not self.current_table:
            return "", []
        
        main = self.escape(self.current_table)
        used = set()
        cols = []
        self.column_mapping = {}
        
        def add(table):
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"PRAGMA table_info({self.escape(table)})")
                for col in cursor.fetchall():
                    name = col[1]
                    if name not in used:
                        sql = f"{self.escape(table)}.{self.escape(name)}"
                        cols.append(sql)
                        self.column_mapping[name] = {'sql': sql, 'table': table, 'name': name}
                        used.add(name)
            except:
                pass
        
        add(self.current_table)
        for j in self.joined_tables:
            add(j['table2'])
        
        if self.selected_attributes:
            final = []
            used.clear()
            self.column_mapping = {}
            for attr in self.selected_attributes:
                if '.' in attr:
                    t, c = attr.split('.')
                    if c not in used:
                        sql = f"{self.escape(t)}.{self.escape(c)}"
                        final.append(sql)
                        self.column_mapping[c] = {'sql': sql, 'table': t, 'name': c}
                        used.add(c)
                else:
                    if attr not in used:
                        sql = self.escape(attr)
                        final.append(sql)
                        self.column_mapping[attr] = {'sql': sql, 'table': self.current_table, 'name': attr}
                        used.add(attr)
            cols = final
        
        if not cols:
            return "", []
        
        select = "SELECT " + ", ".join(cols)
        from_clause = f"FROM {main}"
        joins = []
        for j in self.joined_tables:
            t = self.escape(j['table2'])
            joins.append(f"{j.get('join_type', 'INNER')} JOIN {t} ON {j['condition']}")
        
        order = ""
        if sort_col:
            order = f"ORDER BY {self.escape(sort_col)} {'DESC' if sort_order == 'По убыванию' else 'ASC'}"
        
        query = f"{select} {from_clause} {' '.join(joins)} {order}".strip()
        display = [c.replace('"', '').split('.')[-1] for c in cols]
        return query, display
    
    def isImageColumn(self, name):
        try:
            cursor = self.connection.cursor()
            if self.current_table:
                cursor.execute(f"PRAGMA table_info({self.escape(self.current_table)})")
                for col in cursor.fetchall():
                    if col[1] == name and col[2].upper() == 'BLOB':
                        return True
            for j in self.joined_tables:
                cursor.execute(f"PRAGMA table_info({self.escape(j['table2'])})")
                for col in cursor.fetchall():
                    if col[1] == name and col[2].upper() == 'BLOB':
                        return True
            keywords = ['photo', 'image', 'img', 'picture', 'pic', 'фото']
            return any(k in name.lower() for k in keywords)
        except:
            return False
    
    def isValidImage(self, data):
        if not isinstance(data, bytes) or len(data) < 100:
            return False
        try:
            sig = data[:6]
            return (sig.startswith(b'\xff\xd8\xff') or sig.startswith(b'\x89PNG') or
                    sig.startswith(b'GIF87a') or sig.startswith(b'GIF89a') or sig.startswith(b'BM'))
        except:
            return False
    
    def displayTableData(self, sort_col=None, sort_order="ASC"):
        if not self.current_table and not self.joined_tables:
            return
        
        try:
            self.table.clear()
            query, cols = self.buildQuery(sort_col, sort_order)
            if not cols:
                QMessageBox.warning(self, "Предупреждение", "Нет атрибутов")
                return
            
            cursor = self.connection.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            
            self.table.setRowCount(len(rows))
            self.table.setColumnCount(len(cols))
            self.table.setHorizontalHeaderLabels(cols)
            
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
            self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
            
            self.image_columns = [c for c in cols if self.isImageColumn(c)]
            
            for i, name in enumerate(cols):
                if name in self.image_columns:
                    self.table.setColumnWidth(i, PHOTO_COLUMN_WIDTH)
                    for r in range(len(rows)):
                        self.table.setRowHeight(r, CELL_HEIGHT + 6)
                else:
                    self.table.setColumnWidth(i, TEXT_COLUMN_WIDTH)
            
            for r, row in enumerate(rows):
                for c, val in enumerate(row):
                    name = cols[c]
                    
                    if name in self.image_columns and val and isinstance(val, bytes):
                        if self.isValidImage(val):
                            w = ImageWidget(val, r, c)
                            w.clicked.connect(self.onImageClick)
                            w.rightClicked.connect(self.onImageRightClick)
                            self.table.setCellWidget(r, c, w)
                            item = QTableWidgetItem()
                            item.setData(Qt.ItemDataRole.UserRole, val)
                            self.table.setItem(r, c, item)
                        else:
                            item = QTableWidgetItem("[BLOB]")
                            item.setFont(QFont("Arial", 10))
                            self.table.setItem(r, c, item)
                    elif isinstance(val, bool):
                        item = QTableWidgetItem("✅ Да" if val else "❌ Нет")
                        item.setFont(QFont("Arial", 10))
                        self.table.setItem(r, c, item)
                    elif val is None:
                        item = QTableWidgetItem("")
                        item.setFont(QFont("Arial", 10))
                        self.table.setItem(r, c, item)
                    else:
                        item = QTableWidgetItem(str(val))
                        item.setFont(QFont("Arial", 10))
                        self.table.setItem(r, c, item)
            
            self.sort_col.clear()
            self.sort_col.addItems(self.getAvailableColumns())
            if self.sort_col.count():
                self.sort_col.setCurrentIndex(0)
            
            self.table.resizeRowsToContents()
            self.fixPhotoRowHeights()
                
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def onImageClick(self, r, c):
        name = self.table.horizontalHeaderItem(c).text()
        w = self.table.cellWidget(r, c)
        if w and hasattr(w, 'image_data'):
            self.viewImage(name, w.image_data)
    
    def onImageRightClick(self, r, c):
        menu = QMenu()
        name = self.table.horizontalHeaderItem(c).text()
        
        menu.setFont(QFont("Arial", 10))
        
        add = menu.addAction("📷 Добавить фото")
        view = menu.addAction("🖼️ Просмотреть")
        rem = menu.addAction("🗑️ Удалить")
        
        action = menu.exec(QCursor.pos())
        if action == add:
            self.addPhotoDialog(name, r, c)
        elif action == view:
            self.viewSelectedImage()
        elif action == rem:
            self.removePhoto(r, c, name)
    
    def showContextMenu(self, pos):
        item = self.table.itemAt(pos)
        if not item:
            return
        
        r, c = item.row(), item.column()
        name = self.table.horizontalHeaderItem(c).text()
        
        if self.table.cellWidget(r, c):
            return
        
        has_photo = bool(item.data(Qt.ItemDataRole.UserRole))
        
        if self.isImageColumn(name):
            menu = QMenu()
            menu.setFont(QFont("Arial", 10))
            
            if has_photo:
                view = menu.addAction("🖼️ Просмотреть")
                change = menu.addAction("📷 Изменить")
                rem = menu.addAction("🗑️ Удалить")
            else:
                add = menu.addAction("📷 Добавить")
            
            action = menu.exec(self.table.viewport().mapToGlobal(pos))
            
            if has_photo:
                if action == view:
                    self.viewSelectedImage()
                elif action == change:
                    self.addPhotoDialog(name, r, c)
                elif action == rem:
                    self.removePhoto(r, c, name)
            else:
                if action == add:
                    self.addPhotoDialog(name, r, c)
    
    def onCellDoubleClick(self, idx):
        r, c = idx.row(), idx.column()
        name = self.table.horizontalHeaderItem(c).text()
        
        w = self.table.cellWidget(r, c)
        if w and isinstance(w, ImageWidget):
            self.viewSelectedImage()
            return
        
        item = self.table.item(r, c)
        if not item:
            return
        
        if item.data(Qt.ItemDataRole.UserRole):
            self.addPhotoDialog(name, r, c)
            return
        
        val = item.text()
        info = self.getColumnInfo(name)
        if not info:
            QMessageBox.warning(self, "Ошибка", f"Нет информации о {name}")
            return
        
        t, col = info['table'], info['name']
        typ = self.getColumnType(t, col)
        
        if typ and typ.upper() == 'BOOLEAN':
            dlg = BooleanEditDialog(self, name, val)
            if dlg.exec():
                self.updateCell(r, c, dlg.getValue(), t, col)
        else:
            text, ok = QInputDialog.getText(self, f"Редактирование {name}", "Новое значение:", text=val)
            if ok and text != val:
                self.updateCell(r, c, text, t, col)
    
    def editCell(self):
        item = self.table.currentItem()
        if not item:
            return
        self.onCellDoubleClick(self.table.currentIndex())
    
    def getColumnInfo(self, disp_name):
        clean = disp_name.split('.')[-1] if '.' in disp_name else disp_name
        if clean in self.column_mapping:
            return self.column_mapping[clean]
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape(self.current_table)})")
            for col in cursor.fetchall():
                if col[1] == clean:
                    return {'sql': f"{self.escape(self.current_table)}.{self.escape(clean)}",
                            'table': self.current_table, 'name': clean}
            for j in self.joined_tables:
                cursor.execute(f"PRAGMA table_info({self.escape(j['table2'])})")
                for col in cursor.fetchall():
                    if col[1] == clean:
                        return {'sql': f"{self.escape(j['table2'])}.{self.escape(clean)}",
                                'table': j['table2'], 'name': clean}
        except:
            pass
        return None
    
    def getColumnType(self, table, col):
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape(table)})")
            for c in cursor.fetchall():
                if c[1] == col:
                    return c[2]
        except:
            pass
        return None
    
    def updateCell(self, r, c, new_val, table, col):
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape(table)})")
            pk = cursor.fetchall()[0][1]
            
            pk_idx = -1
            for i in range(self.table.columnCount()):
                h = self.table.horizontalHeaderItem(i).text()
                clean = h.split('.')[-1] if '.' in h else h
                if clean == pk:
                    info = self.getColumnInfo(h)
                    if info and info['table'] == table:
                        pk_idx = i
                        break
            
            if pk_idx == -1:
                QMessageBox.critical(self, "Ошибка", f"Не найден ключ {pk}")
                return
            
            pk_val = self.table.item(r, pk_idx).text()
            
            typ = self.getColumnType(table, col)
            processed = new_val
            if typ and typ.upper() == 'BOOLEAN':
                processed = 1 if str(new_val).lower() in ['true', '1', 'да', 'yes'] else 0
            elif typ and typ.upper() in ['INTEGER', 'REAL']:
                try:
                    processed = float(new_val) if '.' in new_val else int(new_val)
                except:
                    processed = new_val
            
            query = f"UPDATE {self.escape(table)} SET {self.escape(col)} = ? WHERE {pk} = ?"
            cursor.execute(query, (processed, pk_val))
            self.connection.commit()
            
            item = self.table.item(r, c)
            if typ and typ.upper() == 'BOOLEAN':
                item.setText("✅ Да" if processed else "❌ Нет")
            else:
                item.setText(str(new_val))
            
            self.updateStatus(f"✅ Обновлено {table}")
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
            self.connection.rollback()
    
    def getAvailableColumns(self):
        cols = set()
        if self.current_table:
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"PRAGMA table_info({self.escape(self.current_table)})")
                cols.update(c[1] for c in cursor.fetchall())
            except:
                pass
        for j in self.joined_tables:
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"PRAGMA table_info({self.escape(j['table2'])})")
                cols.update(c[1] for c in cursor.fetchall())
            except:
                pass
        return sorted(cols)
    
    def getAllColumns(self):
        all_cols = {}
        used = set()
        
        if self.current_table:
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"PRAGMA table_info({self.escape(self.current_table)})")
                cols = [c[1] for c in cursor.fetchall() if c[1] not in used]
                used.update(cols)
                all_cols[self.current_table] = cols
            except:
                pass
        
        for j in self.joined_tables:
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"PRAGMA table_info({self.escape(j['table2'])})")
                cols = [c[1] for c in cursor.fetchall() if c[1] not in used]
                used.update(cols)
                all_cols[j['table2']] = cols
            except:
                pass
        
        return all_cols
    
    def applySorting(self):
        if (self.current_table or self.joined_tables) and self.sort_col.currentText():
            self.displayTableData(self.sort_col.currentText(), self.sort_order.currentText())
    
    def refreshData(self):
        if self.current_table or self.joined_tables:
            self.displayTableData()
        self.updateTableList()
        self.updateStatus("✅ Обновлено")
    
    def quickAddPhoto(self):
        item = self.table.currentItem()
        if not item:
            QMessageBox.warning(self, "Внимание", "Выберите ячейку")
            return
        
        r, c = item.row(), item.column()
        name = self.table.horizontalHeaderItem(c).text()
        
        if not self.isImageColumn(name):
            reply = QMessageBox.question(self, "Подтверждение",
                                         f"Колонка '{name}' не для фото. Добавить все равно?")
            if reply == QMessageBox.StandardButton.No:
                return
        
        self.addPhotoDialog(name, r, c)
    
    def viewPhoto(self):
        item = self.table.currentItem()
        if not item:
            QMessageBox.warning(self, "Внимание", "Выберите ячейку с фото")
            return
        
        r, c = item.row(), item.column()
        name = self.table.horizontalHeaderItem(c).text()
        
        if name not in self.image_columns:
            QMessageBox.warning(self, "Предупреждение", "Колонка не содержит фото")
            return
        
        w = self.table.cellWidget(r, c)
        if w and hasattr(w, 'image_data'):
            self.viewImage(name, w.image_data)
        elif item.data(Qt.ItemDataRole.UserRole):
            self.viewImage(name, item.data(Qt.ItemDataRole.UserRole))
        else:
            QMessageBox.warning(self, "Предупреждение", "Нет фото")
    
    def addPhotoDialog(self, name, r, c):
        dlg = PhotoDialog(self, name)
        if dlg.exec():
            data = dlg.getImageData()
            if data:
                self.updateImage(r, c, data, name)
    
    def removePhoto(self, r, c, name):
        reply = QMessageBox.question(self, "Удаление", "Удалить фото?")
        if reply == QMessageBox.StandardButton.Yes:
            try:
                cursor = self.connection.cursor()
                cursor.execute(f"PRAGMA table_info({self.escape(self.current_table)})")
                pk = cursor.fetchall()[0][1]
                
                pk_idx = -1
                for i in range(self.table.columnCount()):
                    if self.table.horizontalHeaderItem(i).text() == pk:
                        pk_idx = i
                        break
                
                if pk_idx == -1:
                    QMessageBox.critical(self, "Ошибка", "Не найден ключ")
                    return
                
                pk_val = self.table.item(r, pk_idx).text()
                
                query = f"UPDATE {self.escape(self.current_table)} SET {self.escape(name)} = NULL WHERE {pk} = ?"
                cursor.execute(query, (pk_val,))
                self.connection.commit()
                
                self.table.removeCellWidget(r, c)
                self.table.setItem(r, c, QTableWidgetItem(""))
                
                self.updateStatus("✅ Фото удалено")
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Ошибка", str(e))
    
    def updateImage(self, r, c, data, name):
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape(self.current_table)})")
            pk = cursor.fetchall()[0][1]
            
            pk_idx = -1
            for i in range(self.table.columnCount()):
                if self.table.horizontalHeaderItem(i).text() == pk:
                    pk_idx = i
                    break
            
            if pk_idx == -1:
                QMessageBox.critical(self, "Ошибка", "Не найден ключ")
                return
            
            pk_val = self.table.item(r, pk_idx).text()
            
            query = f"UPDATE {self.escape(self.current_table)} SET {self.escape(name)} = ? WHERE {pk} = ?"
            cursor.execute(query, (data, pk_val))
            self.connection.commit()
            
            self.table.removeCellWidget(r, c)
            w = ImageWidget(data, r, c)
            w.clicked.connect(self.onImageClick)
            w.rightClicked.connect(self.onImageRightClick)
            self.table.setCellWidget(r, c, w)
            
            item = QTableWidgetItem()
            item.setData(Qt.ItemDataRole.UserRole, data)
            self.table.setItem(r, c, item)
            
            self.updateStatus("✅ Фото обновлено")
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def viewSelectedImage(self):
        items = self.table.selectedItems()
        if items:
            r, c = items[0].row(), items[0].column()
        else:
            idx = self.table.currentIndex()
            if not idx.isValid():
                return
            r, c = idx.row(), idx.column()
        
        name = self.table.horizontalHeaderItem(c).text()
        
        if name not in self.image_columns:
            QMessageBox.warning(self, "Предупреждение", "Колонка не содержит фото")
            return
        
        w = self.table.cellWidget(r, c)
        if w and isinstance(w, ImageWidget):
            self.viewImage(name, w.image_data)
            return
        
        item = self.table.item(r, c)
        if item and item.data(Qt.ItemDataRole.UserRole):
            self.viewImage(name, item.data(Qt.ItemDataRole.UserRole))
            return
        
        QMessageBox.warning(self, "Предупреждение", "Нет фото")
    
    def viewImage(self, name, data, info=""):
        ImageViewDialog(self, name, data, info).exec()
    
    def deleteRecord(self):
        items = self.table.selectedItems()
        if not items:
            QMessageBox.warning(self, "Предупреждение", "Выберите запись")
            return
        
        if QMessageBox.question(self, "Подтверждение", "Удалить запись?") != QMessageBox.StandardButton.Yes:
            return
        
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape(self.current_table)})")
            pk = cursor.fetchall()[0][1]
            
            pk_idx = -1
            for i in range(self.table.columnCount()):
                if self.table.horizontalHeaderItem(i).text() == pk:
                    pk_idx = i
                    break
            
            if pk_idx == -1:
                QMessageBox.critical(self, "Ошибка", "Не найден ключ")
                return
            
            pk_val = self.table.item(items[0].row(), pk_idx).text()
            
            cursor.execute(f"DELETE FROM {self.escape(self.current_table)} WHERE {pk} = ?", (pk_val,))
            self.connection.commit()
            
            self.table.removeRow(items[0].row())
            self.updateStatus("✅ Запись удалена")
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def copyCell(self):
        items = self.table.selectedItems()
        if items:
            QApplication.clipboard().setText(items[0].text())
            self.updateStatus("✅ Скопировано")
    
    def copyRow(self):
        items = self.table.selectedItems()
        if items:
            r = items[0].row()
            data = []
            for c in range(self.table.columnCount()):
                if self.table.item(r, c):
                    data.append(self.table.item(r, c).text())
                elif self.table.cellWidget(r, c):
                    data.append("[Фото]")
                else:
                    data.append("")
            QApplication.clipboard().setText("\t".join(data))
            self.updateStatus("✅ Строка скопирована")
    
    def copyHeader(self):
        items = self.table.selectedItems()
        if items:
            h = self.table.horizontalHeaderItem(items[0].column()).text()
            QApplication.clipboard().setText(h)
            self.updateStatus("✅ Заголовок скопирован")
    
    def renameAttribute(self):
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Выберите таблицу")
            return
        
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape(self.current_table)})")
            cols = [c[1] for c in cursor.fetchall()]
            if not cols:
                QMessageBox.warning(self, "Предупреждение", "Нет атрибутов")
                return
            
            old, ok = QInputDialog.getItem(self, "Переименование", "Атрибут:", cols, 0, False)
            if not ok or not old:
                return
            
            new, ok = QInputDialog.getText(self, "Переименование", f"Новое имя для {old}:", text=old)
            if ok and new and new != old:
                self.renameColumn(old, new)
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def renameColumn(self, old, new):
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape(self.current_table)})")
            cols = cursor.fetchall()
            
            new_cols = []
            for col in cols:
                if col[1] == old:
                    new_cols.append(f'"{new}" {col[2]}')
                else:
                    new_cols.append(f'"{col[1]}" {col[2]}')
            
            temp = f"temp_{self.current_table}"
            
            cursor.execute(f"CREATE TABLE {self.escape(temp)} ({', '.join(new_cols)})")
            
            col_names = [f'"{c[1]}"' for c in cols]
            col_list = ', '.join(col_names)
            
            insert_query = f"INSERT INTO {self.escape(temp)} SELECT {col_list} FROM {self.escape(self.current_table)}"
            cursor.execute(insert_query)
            
            cursor.execute(f"DROP TABLE {self.escape(self.current_table)}")
            cursor.execute(f"ALTER TABLE {self.escape(temp)} RENAME TO {self.escape(self.current_table)}")
            
            self.connection.commit()
            self.displayTableData()
            self.updateStatus(f"✅ {old} -> {new}")
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def addColumn(self):
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Выберите таблицу")
            return
        
        dlg = AddColumnDialog(self, self.current_table)
        if dlg.exec():
            name, typ, default = dlg.getData()
            self.addColumnToTable(name, typ, default)
    
    def addColumnToTable(self, name, typ, default=None):
        try:
            cursor = self.connection.cursor()
            query = f"ALTER TABLE {self.escape(self.current_table)} ADD COLUMN {self.escape(name)} {typ}"
            
            if default is not None:
                if typ.upper() == 'BOOLEAN':
                    default = '1' if default.lower() in ['true', '1', 'да'] else '0'
                query += f" DEFAULT {default}"
            
            cursor.execute(query)
            self.connection.commit()
            
            if default is not None:
                cursor.execute(f"UPDATE {self.escape(self.current_table)} SET {self.escape(name)} = ?", (default,))
                self.connection.commit()
            
            self.updateStatus(f"✅ Колонка {name} добавлена")
            self.displayTableData()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def createTable(self):
        dlg = CreateTableDialog(self)
        if dlg.exec():
            name, cols = dlg.getData()
            self.createTableInDB(name, cols)
    
    def createTableInDB(self, name, cols):
        try:
            cursor = self.connection.cursor()
            sql = []
            for col in cols:
                sql.append(f'"{col["name"]}" {col["type"]}')
            query = f"CREATE TABLE IF NOT EXISTS {self.escape(name)} ({', '.join(sql)})"
            cursor.execute(query)
            self.connection.commit()
            self.updateStatus(f"✅ Таблица {name} создана")
            self.updateTableList()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def addRecord(self):
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Выберите таблицу")
            return
        
        dlg = AddRecordDialog(self, self.current_table, self.connection)
        if dlg.exec():
            self.addRecordToTable(dlg.getValues())
    
    def addRecordToTable(self, vals):
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape(self.current_table)})")
            cols = cursor.fetchall()
            names = [c[1] for c in cols]
            types = [c[2] for c in cols]
            
            processed = []
            for i, v in enumerate(vals):
                typ = types[i].upper()
                if v is None or v == "":
                    processed.append(None)
                elif typ == 'BOOLEAN':
                    processed.append(1 if str(v).lower() in ['true', '1', 'да'] else 0)
                else:
                    processed.append(v)
            
            place = ", ".join(["?"] * len(names))
            cursor.execute(f"INSERT INTO {self.escape(self.current_table)} VALUES ({place})", processed)
            self.connection.commit()
            
            self.updateStatus("✅ Запись добавлена")
            self.displayTableData()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def deleteTable(self):
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Выберите таблицу")
            return
        
        if QMessageBox.question(self, "Подтверждение", f"Удалить {self.current_table}?") != QMessageBox.StandardButton.Yes:
            return
        
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"DROP TABLE IF EXISTS {self.escape(self.current_table)}")
            self.connection.commit()
            
            self.updateStatus(f"✅ {self.current_table} удалена")
            self.current_table = None
            self.joined_tables.clear()
            self.selected_attributes.clear()
            if self.current_table in self.table_joins:
                del self.table_joins[self.current_table]
            self.updateTableList()
            self.table.clear()
            self.table.setRowCount(0)
            self.updateJoinInfo()
            self.updateAttributesLabel()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def quickJoin(self):
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Выберите таблицу")
            return
        
        tables = []
        for i in range(self.table_list.count()):
            t = self.table_list.item(i).text()
            if t != self.current_table:
                tables.append(t)
        
        if not tables:
            QMessageBox.information(self, "Информация", "Нет других таблиц")
            return
        
        dlg = MultiTableSelectDialog(self, tables)
        if dlg.exec():
            for t in dlg.getSelectedTables():
                common = self.findCommonColumns(self.current_table, t)
                if not common:
                    QMessageBox.warning(self, "Предупреждение", f"Нет общих полей с {t}")
                    continue
                self.joinTables(t, common[0], common[0])
    
    def findCommonColumns(self, t1, t2):
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape(t1)})")
            c1 = [c[1] for c in cursor.fetchall()]
            cursor.execute(f"PRAGMA table_info({self.escape(t2)})")
            c2 = [c[1] for c in cursor.fetchall()]
            return list(set(c1) & set(c2))
        except:
            return []
    
    def joinTables(self, t2, a1, a2, typ="INNER"):
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape(self.current_table)})")
            if a1 not in [c[1] for c in cursor.fetchall()]:
                QMessageBox.critical(self, "Ошибка", f"{a1} не найден")
                return False
            
            cursor.execute(f"PRAGMA table_info({self.escape(t2)})")
            if a2 not in [c[1] for c in cursor.fetchall()]:
                QMessageBox.critical(self, "Ошибка", f"{a2} не найден")
                return False
            
            for j in self.joined_tables:
                if j['table2'] == t2:
                    QMessageBox.warning(self, "Предупреждение", f"{t2} уже соединена")
                    return False
            
            cond = f"{self.escape(self.current_table)}.{self.escape(a1)} = {self.escape(t2)}.{self.escape(a2)}"
            self.joined_tables.append({'table2': t2, 'condition': cond, 'join_type': typ})
            self.table_joins[self.current_table] = self.joined_tables.copy()
            
            self.updateJoinInfo()
            self.displayTableData()
            self.updateStatus(f"✅ {self.current_table} ↔ {t2}")
            return True
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
            return False
    
    def joinTablesAdvanced(self):
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Выберите таблицу")
            return
        
        dlg = JoinTablesDialog(self, self.current_table, self.connection)
        if dlg.exec():
            t2, a1, a2, typ = dlg.getData()
            self.joinTables(t2, a1, a2, typ)
    
    def selectAttributes(self):
        if not self.current_table and not self.joined_tables:
            QMessageBox.warning(self, "Предупреждение", "Выберите таблицу")
            return
        
        dlg = SelectAttributesDialog(self, self.getAllColumns(), self.selected_attributes)
        if dlg.exec():
            self.selected_attributes = dlg.getSelectedAttributes()
            self.updateAttributesLabel()
            self.displayTableData()
    
    def inspectDB(self):
        if not self.connection:
            QMessageBox.warning(self, "Предупреждение", "Нет подключения")
            return
        
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = cursor.fetchall()
            
            text = "🔍 ИССЛЕДОВАНИЕ\n" + "="*50 + "\n\n"
            text += f"📁 {os.path.basename(self.db_name)}\n"
            text += f"📋 Таблиц: {len(tables)}\n\n"
            
            for t in tables:
                name = t[0]
                text += f"📊 {name}\n" + "-"*30 + "\n"
                
                cursor.execute(f"PRAGMA table_info({self.escape(name)})")
                for col in cursor.fetchall():
                    text += f"  - {col[1]} ({col[2]})\n"
                
                try:
                    cursor.execute(f"SELECT COUNT(*) FROM {self.escape(name)}")
                    text += f"📈 Записей: {cursor.fetchone()[0]}\n"
                except:
                    text += "📈 Записей: -\n"
                text += "\n"
            
            self.showTextDialog("Исследование", text)
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def findAllPhotos(self):
        if not self.connection:
            QMessageBox.warning(self, "Предупреждение", "Нет подключения")
            return
        
        try:
            cursor = self.connection.cursor()
            tables = cursor.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
            
            total = 0
            text = "🖼️ ПОИСК ФОТО\n" + "="*50 + "\n\n"
            
            for t in tables:
                name = t[0]
                text += f"📋 {name}\n"
                
                cursor.execute(f"PRAGMA table_info({self.escape(name)})")
                cols = cursor.fetchall()
                
                found = 0
                for col in cols:
                    if col[2].upper() == 'BLOB' or any(k in col[1].lower() for k in ['photo', 'image', 'pic', 'фото']):
                        text += f"  🔍 {col[1]} ({col[2]})\n"
                        
                        cursor.execute(f"SELECT rowid, {col[1]} FROM {name} WHERE {col[1]} IS NOT NULL")
                        for rowid, data in cursor.fetchall():
                            if isinstance(data, bytes) and len(data) > 100:
                                fname = f"photo_{name}_{col[1]}_{rowid}.jpg"
                                try:
                                    with open(fname, 'wb') as f:
                                        f.write(data)
                                    text += f"    ✅ {fname} ({len(data)} bytes)\n"
                                    total += 1
                                    found += 1
                                except Exception as e:
                                    text += f"    ❌ {e}\n"
                
                if found:
                    text += f"  📊 Найдено: {found}\n"
                else:
                    text += "  ❌ Нет фото\n"
                text += "\n"
            
            if total:
                text += f"✅ Всего: {total}\n"
            else:
                text += "⚠ Фото не найдены\n"
            
            self.showTextDialog("Результаты", text)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def showTextDialog(self, title, text):
        dlg = QDialog(self)
        dlg.setWindowTitle(title)
        dlg.setGeometry(100, 100, 800, 600)
        
        layout = QVBoxLayout(dlg)
        
        edit = QTextEdit()
        edit.setPlainText(text)
        edit.setReadOnly(True)
        edit.setFont(QFont("Consolas", 10))
        
        btns = QHBoxLayout()
        save = QPushButton("💾 Сохранить")
        save.clicked.connect(lambda: self.saveText(text, title))
        close = QPushButton("❌ Закрыть")
        close.clicked.connect(dlg.close)
        
        btns.addWidget(save)
        btns.addStretch()
        btns.addWidget(close)
        
        layout.addWidget(edit)
        layout.addLayout(btns)
        applyTextFit(dlg)
        dlg.setMinimumSize(550, 450)
        dlg.exec()
    
    def saveText(self, text, title):
        path, _ = QFileDialog.getSaveFileName(self, f"Сохранить {title}", "", "Text files (*.txt)")
        if path:
            try:
                with open(path, 'w', encoding='utf-8') as f:
                    f.write(text)
                self.updateStatus(f"✅ Сохранено {os.path.basename(path)}")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))
    
    def importExcel(self):
        if not self.current_table:
            QMessageBox.warning(self, "Предупреждение", "Выберите таблицу")
            return
        
        path, _ = QFileDialog.getOpenFileName(self, "Выберите Excel", "", "Excel files (*.xlsx *.xls)")
        if not path:
            return
        
        try:
            df = pd.read_excel(path)
            if df.empty:
                QMessageBox.warning(self, "Предупреждение", "Файл пуст")
                return
            
            dlg = ExcelImportDialog(self, df.columns.tolist())
            if not dlg.exec():
                return
            
            cursor = self.connection.cursor()
            cursor.execute(f"PRAGMA table_info({self.escape(self.current_table)})")
            cols = [c[1] for c in cursor.fetchall()]
            
            for _, row in df.iterrows():
                vals = []
                for c in cols:
                    if c in df.columns:
                        v = row[c]
                        vals.append(None if pd.isna(v) else v)
                    else:
                        vals.append(None)
                
                place = ", ".join(["?"] * len(cols))
                cursor.execute(f"INSERT INTO {self.escape(self.current_table)} VALUES ({place})", vals)
            
            self.connection.commit()
            self.displayTableData()
            self.updateStatus(f"✅ Импортировано из {os.path.basename(path)}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def exportExcelWithPhotos(self):
        if not self.current_table and not self.joined_tables:
            QMessageBox.warning(self, "Предупреждение", "Нет данных")
            return
        
        dlg = ExportSettingsDialog(self)
        if not dlg.exec():
            return
        
        settings = dlg.getSettings()
        
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить Excel", "", "Excel files (*.xlsx)")
        if not path:
            return
        
        try:
            query, cols = self.buildQuery()
            cursor = self.connection.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as ExcelImage
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = self.current_table or "Данные"
            
            for i, name in enumerate(cols, 1):
                ws.cell(row=1, column=i, value=name)
                ws.column_dimensions[get_column_letter(i)].width = 15
            
            photo_count = 0
            saved_files = []
            temp_dir = tempfile.mkdtemp(prefix="excel_")
            temp_files = []
            
            try:
                for r, row in enumerate(rows, 2):
                    for c, val in enumerate(row, 1):
                        name = cols[c-1]
                        
                        if (name in self.image_columns and val and isinstance(val, bytes) and 
                            settings['include_images'] and self.isValidImage(val)):
                            
                            temp = os.path.join(temp_dir, f"photo_{r}_{c}.png")
                            with open(temp, 'wb') as f:
                                f.write(val)
                            temp_files.append(temp)
                            
                            if settings['save_as_files']:
                                save_dir = os.path.dirname(path) or "."
                                fname = f"{self.current_table}_row{r-1}_{name}.png"
                                fpath = os.path.join(save_dir, fname)
                                os.makedirs(save_dir, exist_ok=True)
                                shutil.copy2(temp, fpath)
                                saved_files.append(fpath)
                                ws.cell(row=r, column=c, value=f"📷 {fname}")
                            else:
                                try:
                                    img = ExcelImage(temp)
                                    img.width = settings['image_size']
                                    img.height = settings['image_size']
                                    ws.add_image(img, f"{get_column_letter(c)}{r}")
                                    ws.row_dimensions[r].height = settings['image_size'] * 0.75
                                    photo_count += 1
                                except:
                                    ws.cell(row=r, column=c, value="[Фото]")
                        
                        elif name in self.image_columns and val:
                            ws.cell(row=r, column=c, value="🖼️ Фото")
                        elif isinstance(val, bool):
                            ws.cell(row=r, column=c, value="✅ Да" if val else "❌ Нет")
                        elif val is None:
                            ws.cell(row=r, column=c, value="")
                        else:
                            ws.cell(row=r, column=c, value=str(val))
                
                info = wb.create_sheet(title="Информация")
                info['A1'] = "Отчет"
                info['A3'] = f"Таблица: {self.current_table}"
                info['A4'] = f"База: {os.path.basename(self.db_name)}"
                info['A5'] = f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                info['A6'] = f"Строк: {len(rows)}"
                info['A7'] = f"Колонок: {len(cols)}"
                info['A8'] = f"Фото: {photo_count}"
                
                if saved_files:
                    info['A10'] = "Сохраненные фото:"
                    for i, f in enumerate(saved_files, 11):
                        info[f'A{i}'] = os.path.basename(f)
                
                wb.save(path)
                
                report = f"✅ Экспорт завершен\n\nФайл: {os.path.basename(path)}\nСтрок: {len(rows)}\nКолонок: {len(cols)}"
                if settings['include_images']:
                    if settings['save_as_files']:
                        report += f"\nФото файлов: {len(saved_files)}"
                    else:
                        report += f"\nФото в Excel: {photo_count}"
                
                self.updateStatus(f"✅ Экспорт {os.path.basename(path)}")
                QMessageBox.information(self, "Успех", report)
            
            finally:
                for f in temp_files:
                    try:
                        if os.path.exists(f):
                            os.unlink(f)
                    except:
                        pass
                try:
                    if os.path.exists(temp_dir):
                        os.rmdir(temp_dir)
                except:
                    pass
                    
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def printData(self):
        """Полноценная печать данных в PDF с авто-подбором высоты строк и переносом текста"""
        if not self.current_table and not self.joined_tables:
            QMessageBox.warning(self, "Предупреждение", "Нет данных")
            return
        
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить PDF", "", "PDF files (*.pdf)")
        if not path:
            return
        
        try:
            query, cols = self.buildQuery()
            cursor = self.connection.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            
            if not rows:
                QMessageBox.information(self, "Информация", "Нет данных")
                return
            
            pdf = canvas.Canvas(path, pagesize=landscape(A4))
            pdf.setTitle(f"База - {self.current_table}")
            
            # Шрифт
            if self.russian_font_registered:
                pdf.setFont('RussianFont', 10)
            else:
                pdf.setFont('Helvetica', 10)
            
            # Параметры
            margin = 40
            page_width, page_height = landscape(A4)
            table_width = page_width - 2 * margin
            
            # Рассчитываем ширину колонок (поровну)
            num_cols = len(cols)
            col_width = table_width / num_cols
            if col_width < 40: col_width = 40  # минимальная ширина
            
            y = page_height - margin
            
            # --- ЗАГОЛОВОК ОТЧЕТА ---
            pdf.setFontSize(18)
            pdf.setFont('RussianFont' if self.russian_font_registered else 'Helvetica-Bold', 18)
            pdf.drawString(margin, y, f"Отчет: {self.current_table}")
            y -= 30
            
            pdf.setFontSize(10)
            pdf.setFont('RussianFont' if self.russian_font_registered else 'Helvetica', 10)
            pdf.drawString(margin, y, f"База данных: {os.path.basename(self.db_name)}")
            y -= 18
            pdf.drawString(margin, y, f"Дата создания: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            y -= 30
            
            # --- ПОДГОТОВКА ДАННЫХ ДЛЯ ОТРИСОВКИ ---
            # Чтобы не пересчитывать много раз, подготовим всё заранее
            formatted_rows = []
            img_cols = [c for c in cols if self.isImageColumn(c)]
            
            for r, row in enumerate(rows):
                formatted_row = []
                max_height = 25  # Минимальная высота строки
                
                for c, val in enumerate(row):
                    name = cols[c]
                    cell_data = {'text_lines': [], 'image_data': None, 'height': 25}
                    
                    if name in img_cols and val and isinstance(val, bytes) and self.isValidImage(val):
                        cell_data['image_data'] = val
                        cell_data['height'] = 100  # Фото требует высоты
                    
                    elif val is not None and not isinstance(val, bool):
                        text = str(val)
                        # Перенос текста
                        lines = self.wrap_text(text, int(col_width / 5.5))
                        cell_data['text_lines'] = lines
                        # Высота строки = количество строк * 12px + запас
                        h = len(lines) * 12 + 8
                        if h > cell_data['height']:
                            cell_data['height'] = h
                    
                    else:
                        # Булевы значения и пустые
                        if isinstance(val, bool):
                            cell_data['text_lines'] = ["✅ Да" if val else "❌ Нет"]
                        else:
                            cell_data['text_lines'] = [""]
                        cell_data['height'] = 20
                    
                    if cell_data['height'] > max_height:
                        max_height = cell_data['height']
                    
                    formatted_row.append(cell_data)
                
                # Чтобы все ячейки в строке были одной высоты
                for cell in formatted_row:
                    cell['height'] = max_height
                
                formatted_rows.append({'cells': formatted_row, 'height': max_height})
            
            # --- ОТРИСОВКА ТАБЛИЦЫ ---
            img_temp_files = []
            
            # 1. Рисуем заголовки колонок
            header_height = 25
            x = margin
            pdf.setFontSize(9)
            pdf.setFont('RussianFont' if self.russian_font_registered else 'Helvetica-Bold', 9)
            
            # Верхняя линия заголовка
            pdf.line(margin, y, margin + table_width, y)
            y -= header_height
            pdf.line(margin, y, margin + table_width, y)
            
            for i, name in enumerate(cols):
                # Вертикальные линии заголовка
                pdf.line(x, y + header_height, x, y)
                pdf.drawString(x + 4, y + 8, str(name))
                x += col_width
            pdf.line(margin + table_width, y + header_height, margin + table_width, y) # Правая линия
            
            # 2. Рисуем строки данных
            pdf.setFontSize(8)
            pdf.setFont('RussianFont' if self.russian_font_registered else 'Helvetica', 8)
            
            # Инициализируем временные файлы для фотографий перед циклом, чтобы не плодить их
            import tempfile
            
            for row_data in formatted_rows:
                row_height = row_data['height'] + 6 # +6 на отступы
                
                # Если не хватает места, создаем новую страницу
                if y - row_height < margin:
                    pdf.showPage()
                    if self.russian_font_registered:
                        pdf.setFont('RussianFont', 12)
                    else:
                        pdf.setFont('Helvetica', 12)
                    pdf.drawString(margin, page_height - 40, f"{self.current_table} (продолжение)")
                    y = page_height - margin - 20
                    # Перерисовываем заголовки
                    x = margin
                    pdf.setFontSize(8)
                    pdf.setFont('RussianFont-Bold' if self.russian_font_registered else 'Helvetica-Bold', 8)
                    pdf.line(margin, y, margin + table_width, y)
                    y -= header_height
                    pdf.line(margin, y, margin + table_width, y)
                    for i, name in enumerate(cols):
                        pdf.line(x, y + header_height, x, y)
                        pdf.drawString(x + 4, y + 8, str(name))
                        x += col_width
                    pdf.line(margin + table_width, y + header_height, margin + table_width, y)
                    y -= 4 # маленький отступ
                
                # Рисуем нижнюю линию для этой строки
                y -= row_height
                pdf.line(margin, y, margin + table_width, y)
                
                x = margin
                for i, cell in enumerate(row_data['cells']):
                    # Вертикальные линии
                    pdf.line(x, y + row_height, x, y)
                    
                    if cell['image_data']:
                        # ВСТАВКА ФОТО
                        try:
                            temp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                            temp.write(cell['image_data'])
                            temp.close()
                            img_temp_files.append(temp.name)
                            
                            max_w = col_width - 8
                            max_h = row_height - 8
                            img = Image.open(BytesIO(cell['image_data']))
                            img_w, img_h = img.size
                            ratio = min(max_w / img_w, max_h / img_h)
                            new_w = int(img_w * ratio)
                            new_h = int(img_h * ratio)
                            
                            if ratio < 1:
                                img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
                            img.save(temp.name, format='PNG')
                            
                            # Центрируем
                            img_x = x + (col_width - new_w) / 2
                            img_y = y + (row_height - new_h) / 2
                            pdf.drawImage(ImageReader(temp.name), img_x, img_y,
                                         width=new_w, height=new_h, preserveAspectRatio=True)
                        except Exception as e:
                            pdf.drawString(x + 4, y + row_height/2 - 4, "⚠️ Ошибка фото")
                    
                    elif cell['text_lines']:
                        # ВСТАВКА ТЕКСТА С ПЕРЕНОСОМ
                        lines = cell['text_lines']
                        line_height = 11
                        # Вертикальное центрирование многострочного текста
                        total_text_height = len(lines) * line_height
                        start_y = y + (row_height - total_text_height) / 2 + (line_height - 2)
                        
                        for j, line in enumerate(lines):
                            pdf.drawString(x + 4, start_y - j * line_height, line)
                    
                    x += col_width
                
                # Правая граница строки
                pdf.line(margin + table_width, y + row_height, margin + table_width, y)
            
            pdf.save()
            
            # Очистка временных файлов
            for f in img_temp_files:
                try:
                    if os.path.exists(f):
                        os.unlink(f)
                except:
                    pass
            
            self.updateStatus(f"✅ PDF сохранен: {os.path.basename(path)}")
            QMessageBox.information(self, "Успех", f"PDF успешно создан и сохранен:\n{path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))
            print(f"Ошибка при создании PDF: {e}")
    
    def updateStatus(self, msg):
        self.status.showMessage(msg)
        QTimer.singleShot(3000, lambda: self.status.showMessage("✅ Готов к работе"))


# Диалоги
class BooleanEditDialog(QDialog):
    def __init__(self, parent, name, cur):
        super().__init__(parent)
        self.setWindowTitle(f"Редактирование {name}")
        self.setGeometry(300, 300, 300, 150)
        self.setFont(QFont("Arial", 10))
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f"Значение для {name}:"))
        
        self.group = QButtonGroup(self)
        true_btn = QRadioButton("✅ Да")
        false_btn = QRadioButton("❌ Нет")
        
        cur_bool = str(cur).lower() in ['1', 'true', 'да', '✅ да']
        true_btn.setChecked(cur_bool)
        false_btn.setChecked(not cur_bool)
        
        self.group.addButton(true_btn, 1)
        self.group.addButton(false_btn, 0)
        
        radio = QHBoxLayout()
        radio.addWidget(true_btn)
        radio.addWidget(false_btn)
        layout.addLayout(radio)
        
        btns = QHBoxLayout()
        ok_btn = QPushButton("✅ OK")
        ok_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("❌ Отмена")
        cancel_btn.clicked.connect(self.reject)
        btns.addWidget(ok_btn)
        btns.addWidget(cancel_btn)
        layout.addLayout(btns)
        applyTextFit(self)
        self.setMinimumSize(320, 180)
    
    def getValue(self):
        return "True" if self.group.checkedId() == 1 else "False"


class PhotoDialog(QDialog):
    def __init__(self, parent, name):
        super().__init__(parent)
        self.setWindowTitle(f"Добавить фото - {name}")
        self.setGeometry(300, 300, 500, 400)
        self.setFont(QFont("Arial", 10))
        
        self.data = None
        layout = QVBoxLayout(self)
        
        layout.addWidget(QLabel("📸 Добавление фотографии"))
        
        self.preview = QLabel("Выберите изображение")
        self.preview.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.preview.setMinimumHeight(200)
        layout.addWidget(self.preview)
        
        self.info = QLabel("")
        layout.addWidget(self.info)
        
        btns = QHBoxLayout()
        select = QPushButton("📁 Выбрать")
        select.clicked.connect(self.loadImage)
        save = QPushButton("✅ Сохранить")
        save.clicked.connect(self.accept)
        cancel = QPushButton("❌ Отмена")
        cancel.clicked.connect(self.reject)
        
        btns.addWidget(select)
        btns.addWidget(save)
        btns.addWidget(cancel)
        layout.addLayout(btns)
        
        tips = QLabel("💡 PNG, JPG, GIF, BMP | до 5 МБ")
        tips.setStyleSheet("color: gray; font-size: 10px;")
        layout.addWidget(tips)
        applyTextFit(self)
        self.setMinimumSize(420, 380)
    
    def loadImage(self):
        path, _ = QFileDialog.getOpenFileName(self, "Выберите изображение", "",
                                              "Images (*.png *.jpg *.jpeg *.gif *.bmp)")
        if path:
            try:
                with open(path, 'rb') as f:
                    self.data = f.read()
                
                pix = QPixmap(path)
                if not pix.isNull():
                    self.preview.setPixmap(
                        pix.scaled(
                            300,
                            300,
                            Qt.AspectRatioMode.KeepAspectRatio,
                            Qt.TransformationMode.SmoothTransformation,
                        )
                    )
                    self.info.setText(f"Файл: {os.path.basename(path)}\nРазмер: {len(self.data)} байт")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))
    
    def getImageData(self):
        return self.data


class ImageViewDialog(QDialog):
    def __init__(self, parent, name, data, info=""):
        super().__init__(parent)
        self.setWindowTitle(f"Фото - {name} {info}")
        self.setGeometry(100, 100, 900, 700)
        self.setFont(QFont("Arial", 10))
        
        self.data = data
        self.pixmap = None
        self._qimage_buffer = None
        self.scale = 1.0
        self.min_scale = 0.1
        self.max_scale = 3.0
        
        layout = QVBoxLayout(self)
        
        # Панель масштабирования
        zoom = QHBoxLayout()
        
        zoom_out = QPushButton("−")
        zoom_out.setFixedSize(30, 30)
        zoom_out.clicked.connect(self.zoomOut)
        
        self.slider = QSlider(Qt.Orientation.Horizontal)
        self.slider.setRange(10, 300)
        self.slider.setValue(100)
        self.slider.valueChanged.connect(self.sliderChanged)
        
        zoom_in = QPushButton("+")
        zoom_in.setFixedSize(30, 30)
        zoom_in.clicked.connect(self.zoomIn)
        
        self.percent = QLabel("100%")
        self.percent.setMinimumWidth(50)
        self.percent.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        reset = QPushButton("100%")
        reset.clicked.connect(self.resetZoom)
        
        zoom.addWidget(zoom_out)
        zoom.addWidget(self.slider)
        zoom.addWidget(zoom_in)
        zoom.addWidget(self.percent)
        zoom.addWidget(reset)
        zoom.addStretch()
        
        layout.addLayout(zoom)
        
        # Изображение
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.label = QLabel()
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.label.setScaledContents(False)
        
        scroll.setWidget(self.label)
        layout.addWidget(scroll)
        
        self.loadImage()
        
        info_label = QLabel(f"Размер: {self.orig_w}x{self.orig_h} | Объем: {len(data)} байт")
        layout.addWidget(info_label)
        
        # Кнопки
        btns = QHBoxLayout()
        save = QPushButton("💾 Сохранить")
        save.clicked.connect(self.saveImage)
        print_btn = QPushButton("🖨️ Печать")
        print_btn.clicked.connect(self.printImage)
        close = QPushButton("❌ Закрыть")
        close.clicked.connect(self.accept)
        
        btns.addWidget(save)
        btns.addWidget(print_btn)
        btns.addStretch()
        btns.addWidget(close)
        layout.addLayout(btns)
        applyTextFit(self)
        self.setMinimumSize(600, 500)
        self.setupHotkeys()
    
    def loadImage(self):
        try:
            img = Image.open(BytesIO(self.data))
            self.orig_w, self.orig_h = img.size
            
            if img.mode == 'RGBA':
                self._qimage_buffer = img.tobytes()
                qimg = QImage(self._qimage_buffer, img.width, img.height, img.width * 4, QImage.Format.Format_RGBA8888)
            else:
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                self._qimage_buffer = img.tobytes()
                qimg = QImage(self._qimage_buffer, img.width, img.height, img.width * 3, QImage.Format.Format_RGB888)
            
            self.pixmap = QPixmap.fromImage(qimg)
            self.updateScale()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def updateScale(self):
        if self.pixmap:
            dpr = self.label.devicePixelRatioF()
            w = max(1, int(self.pixmap.width() * self.scale))
            h = max(1, int(self.pixmap.height() * self.scale))
            scaled = self.pixmap.scaled(max(1, int(w * dpr)), max(1, int(h * dpr)),
                                        Qt.AspectRatioMode.KeepAspectRatio,
                                        Qt.TransformationMode.SmoothTransformation)
            scaled.setDevicePixelRatio(dpr)
            self.label.setPixmap(scaled)
            self.percent.setText(f"{int(self.scale * 100)}%")
            self.slider.blockSignals(True)
            self.slider.setValue(int(self.scale * 100))
            self.slider.blockSignals(False)
    
    def zoomIn(self):
        if self.scale < self.max_scale:
            self.scale = min(self.scale + 0.1, self.max_scale)
            self.updateScale()
    
    def zoomOut(self):
        if self.scale > self.min_scale:
            self.scale = max(self.scale - 0.1, self.min_scale)
            self.updateScale()
    
    def resetZoom(self):
        self.scale = 1.0
        self.updateScale()
    
    def sliderChanged(self, val):
        self.scale = val / 100.0
        self.updateScale()
    
    def setupHotkeys(self):
        QShortcut(QKeySequence("Ctrl++"), self).activated.connect(self.zoomIn)
        QShortcut(QKeySequence("Ctrl+-"), self).activated.connect(self.zoomOut)
        QShortcut(QKeySequence("Ctrl+0"), self).activated.connect(self.resetZoom)
    
    def saveImage(self):
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить", "", "PNG files (*.png);;JPEG files (*.jpg)")
        if path:
            try:
                with open(path, 'wb') as f:
                    f.write(self.data)
                QMessageBox.information(self, "Успех", f"Сохранено:\n{path}")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))
    
    def printImage(self):
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        dlg = QPrintDialog(printer, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            painter = QPainter(printer)
            pix = self.label.pixmap()
            if pix:
                rect = printer.pageRect(QPrinter.Unit.DevicePixel)
                scaled = pix.scaled(rect.width(), rect.height(),
                                    Qt.AspectRatioMode.KeepAspectRatio,
                                    Qt.TransformationMode.SmoothTransformation)
                x = (rect.width() - scaled.width()) / 2
                y = (rect.height() - scaled.height()) / 2
                painter.drawPixmap(int(x), int(y), scaled)
            painter.end()


class AddColumnDialog(QDialog):
    def __init__(self, parent, table):
        super().__init__(parent)
        self.setWindowTitle("Добавить колонку")
        self.setGeometry(300, 300, 400, 300)
        self.setFont(QFont("Arial", 10))
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f"Колонка в '{table}'"))
        
        layout.addWidget(QLabel("Имя:"))
        self.name_edit = QLineEdit()
        self.name_edit.setFont(QFont("Arial", 10))
        layout.addWidget(self.name_edit)
        
        layout.addWidget(QLabel("Тип:"))
        self.type_combo = QComboBox()
        self.type_combo.setFont(QFont("Arial", 10))
        self.type_combo.addItems(["TEXT", "INTEGER", "REAL", "BOOLEAN", "BLOB"])
        layout.addWidget(self.type_combo)
        
        layout.addWidget(QLabel("По умолчанию (необязательно):"))
        self.default_edit = QLineEdit()
        self.default_edit.setFont(QFont("Arial", 10))
        layout.addWidget(self.default_edit)
        
        help_text = "💡 TEXT - текст\n💡 INTEGER - целые\n💡 REAL - дробные\n💡 BOOLEAN - да/нет\n💡 BLOB - фото"
        help_label = QLabel(help_text)
        help_label.setStyleSheet("color: gray; font-size: 10px;")
        layout.addWidget(help_label)
        
        btns = QHBoxLayout()
        add = QPushButton("✅ Добавить")
        add.clicked.connect(self.accept)
        cancel = QPushButton("❌ Отмена")
        cancel.clicked.connect(self.reject)
        btns.addWidget(add)
        btns.addWidget(cancel)
        layout.addLayout(btns)
        applyTextFit(self)
        self.setMinimumSize(380, 320)
    
    def getData(self):
        name = self.name_edit.text().strip()
        typ = self.type_combo.currentText()
        default = self.default_edit.text().strip()
        return name, typ, default if default else None


class MultiTableSelectDialog(QDialog):
    def __init__(self, parent, tables):
        super().__init__(parent)
        self.setWindowTitle("Выбор таблиц")
        self.setGeometry(300, 300, 400, 500)
        self.setFont(QFont("Arial", 10))
        
        self.tables = tables
        self.checkboxes = []
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("🔗 Выберите таблицы:"))
        
        for t in tables:
            cb = QCheckBox(t)
            cb.setFont(QFont("Arial", 10))
            self.checkboxes.append(cb)
            layout.addWidget(cb)
        
        layout.addStretch()
        
        btns = QHBoxLayout()
        select = QPushButton("✅ Выбрать все")
        select.clicked.connect(self.selectAll)
        deselect = QPushButton("❌ Снять все")
        deselect.clicked.connect(self.deselectAll)
        btns.addWidget(select)
        btns.addWidget(deselect)
        layout.addLayout(btns)
        
        info = QLabel("ℹ️ Авто-соединение по общим полям")
        info.setStyleSheet("color: gray; font-size: 10px;")
        layout.addWidget(info)
        
        dlg_btns = QHBoxLayout()
        join = QPushButton("🔗 Соединить")
        join.clicked.connect(self.accept)
        cancel = QPushButton("❌ Отмена")
        cancel.clicked.connect(self.reject)
        dlg_btns.addWidget(join)
        dlg_btns.addWidget(cancel)
        layout.addLayout(dlg_btns)
        applyTextFit(self)
        self.setMinimumSize(380, 400)
    
    def selectAll(self):
        for cb in self.checkboxes:
            cb.setChecked(True)
    
    def deselectAll(self):
        for cb in self.checkboxes:
            cb.setChecked(False)
    
    def getSelectedTables(self):
        return [self.tables[i] for i, cb in enumerate(self.checkboxes) if cb.isChecked()]


class ExcelImportDialog(QDialog):
    def __init__(self, parent, cols):
        super().__init__(parent)
        self.setWindowTitle("Импорт Excel")
        self.setGeometry(300, 300, 500, 400)
        self.setFont(QFont("Arial", 10))
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("📥 Импорт из Excel"))
        layout.addWidget(QLabel(f"Колонок в Excel: {len(cols)}"))
        
        warn = QLabel("⚠️ Структура должна совпадать")
        warn.setStyleSheet("color: orange; font-size: 10px;")
        layout.addWidget(warn)
        
        layout.addWidget(QLabel("Колонки в файле:"))
        
        list_widget = QListWidget()
        list_widget.setFont(QFont("Arial", 10))
        for c in cols:
            list_widget.addItem(c)
        layout.addWidget(list_widget)
        
        btns = QHBoxLayout()
        imp = QPushButton("✅ Импортировать")
        imp.clicked.connect(self.accept)
        cancel = QPushButton("❌ Отмена")
        cancel.clicked.connect(self.reject)
        btns.addWidget(imp)
        btns.addWidget(cancel)
        layout.addLayout(btns)
        applyTextFit(self)
        self.setMinimumSize(450, 420)


class JoinTablesDialog(QDialog):
    def __init__(self, parent, table, conn):
        super().__init__(parent)
        self.setWindowTitle("Соединить таблицы")
        self.setGeometry(300, 300, 500, 400)
        self.setFont(QFont("Arial", 10))
        
        self.table = table
        self.conn = conn
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f"🔗 Основная: {table}"))
        
        layout.addWidget(QLabel("Таблица для соединения:"))
        self.table2 = QComboBox()
        self.table2.setFont(QFont("Arial", 10))
        self.loadTables()
        layout.addWidget(self.table2)
        
        layout.addWidget(QLabel("Атрибут из основной:"))
        self.attr1 = QComboBox()
        self.attr1.setFont(QFont("Arial", 10))
        self.loadAttributes(table, self.attr1)
        layout.addWidget(self.attr1)
        
        layout.addWidget(QLabel("Атрибут из второй:"))
        self.attr2 = QComboBox()
        self.attr2.setFont(QFont("Arial", 10))
        layout.addWidget(self.attr2)
        
        layout.addWidget(QLabel("Тип соединения:"))
        self.join_type = QComboBox()
        self.join_type.setFont(QFont("Arial", 10))
        self.join_type.addItems(["INNER JOIN", "LEFT JOIN"])
        layout.addWidget(self.join_type)
        
        layout.addWidget(QLabel("Предпросмотр:"))
        self.preview = QTextEdit()
        self.preview.setFont(QFont("Arial", 10))
        self.preview.setReadOnly(True)
        self.preview.setMaximumHeight(100)
        layout.addWidget(self.preview)
        
        self.table2.currentTextChanged.connect(self.updateAttr2)
        self.attr1.currentTextChanged.connect(self.updatePreview)
        self.attr2.currentTextChanged.connect(self.updatePreview)
        self.join_type.currentTextChanged.connect(self.updatePreview)
        
        btns = QHBoxLayout()
        join = QPushButton("🔗 Соединить")
        join.clicked.connect(self.accept)
        cancel = QPushButton("❌ Отмена")
        cancel.clicked.connect(self.reject)
        btns.addWidget(join)
        btns.addWidget(cancel)
        layout.addLayout(btns)
        applyTextFit(self)
        self.setMinimumSize(450, 420)
        self.updatePreview()
    
    def loadTables(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            for t in cursor.fetchall():
                if t[0] != self.table and t[0] != "sqlite_sequence":
                    self.table2.addItem(t[0])
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def loadAttributes(self, table, combo):
        try:
            cursor = self.conn.cursor()
            cursor.execute(f"PRAGMA table_info('{table}')")
            combo.clear()
            for c in cursor.fetchall():
                combo.addItem(c[1])
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def updateAttr2(self):
        t = self.table2.currentText()
        if t:
            self.loadAttributes(t, self.attr2)
            self.updatePreview()
    
    def updatePreview(self):
        t2 = self.table2.currentText()
        a1 = self.attr1.currentText()
        a2 = self.attr2.currentText()
        typ = self.join_type.currentText().split()[0]
        
        if t2 and a1 and a2:
            q = f"SELECT *\nFROM {self.table}\n{typ} JOIN {t2}\nON {self.table}.{a1} = {t2}.{a2}"
            self.preview.setText(q)
    
    def getData(self):
        return (self.table2.currentText(), self.attr1.currentText(),
                self.attr2.currentText(), self.join_type.currentText().split()[0])


class SelectAttributesDialog(QDialog):
    def __init__(self, parent, all_cols, selected):
        super().__init__(parent)
        self.setWindowTitle("Выбор атрибутов")
        self.setGeometry(300, 300, 500, 600)
        self.setFont(QFont("Arial", 10))
        
        self.all_cols = all_cols
        self.selected = selected.copy()
        self.checkboxes = {}
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("👁️ Выберите атрибуты:"))
        
        scroll = QScrollArea()
        widget = QWidget()
        scroll_layout = QVBoxLayout(widget)
        
        for table, cols in all_cols.items():
            scroll_layout.addWidget(QLabel(f"📋 {table}"))
            for c in cols:
                full = f"{table}.{c}"
                cb = QCheckBox(c)
                cb.setFont(QFont("Arial", 10))
                cb.setChecked(full in selected)
                self.checkboxes[full] = cb
                scroll_layout.addWidget(cb)
            scroll_layout.addSpacing(10)
        
        scroll.setWidget(widget)
        layout.addWidget(scroll)
        
        btns = QHBoxLayout()
        select = QPushButton("✅ Выбрать все")
        select.clicked.connect(self.selectAll)
        deselect = QPushButton("❌ Снять все")
        deselect.clicked.connect(self.deselectAll)
        btns.addWidget(select)
        btns.addWidget(deselect)
        layout.addLayout(btns)
        
        dlg_btns = QHBoxLayout()
        apply = QPushButton("✅ Применить")
        apply.clicked.connect(self.accept)
        cancel = QPushButton("❌ Отмена")
        cancel.clicked.connect(self.reject)
        show_all = QPushButton("👁️ Показать все")
        show_all.clicked.connect(self.showAll)
        dlg_btns.addWidget(apply)
        dlg_btns.addWidget(cancel)
        dlg_btns.addWidget(show_all)
        layout.addLayout(dlg_btns)
        applyTextFit(self)
        self.setMinimumSize(450, 500)
    
    def selectAll(self):
        for cb in self.checkboxes.values():
            cb.setChecked(True)
    
    def deselectAll(self):
        for cb in self.checkboxes.values():
            cb.setChecked(False)
    
    def showAll(self):
        self.selected = []
        self.accept()
    
    def getSelectedAttributes(self):
        return [name for name, cb in self.checkboxes.items() if cb.isChecked()]


class CreateTableDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Создать таблицу")
        self.setGeometry(300, 300, 600, 500)
        self.setFont(QFont("Arial", 10))
        
        self.columns = []
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("📊 Новая таблица"))
        
        layout.addWidget(QLabel("Название:"))
        self.name_edit = QLineEdit()
        self.name_edit.setFont(QFont("Arial", 10))
        layout.addWidget(self.name_edit)
        
        group = QGroupBox("📋 Колонки")
        group.setFont(QFont("Arial", 10))
        group_layout = QVBoxLayout()
        
        self.list = QListWidget()
        self.list.setFont(QFont("Arial", 10))
        group_layout.addWidget(self.list)
        
        btn_layout = QHBoxLayout()
        add = QPushButton("➕ Добавить")
        add.clicked.connect(self.addColumn)
        remove = QPushButton("🗑️ Удалить")
        remove.clicked.connect(self.removeColumn)
        btn_layout.addWidget(add)
        btn_layout.addWidget(remove)
        group_layout.addLayout(btn_layout)
        
        group.setLayout(group_layout)
        layout.addWidget(group)
        
        dlg_btns = QHBoxLayout()
        create = QPushButton("✅ Создать")
        create.clicked.connect(self.accept)
        cancel = QPushButton("❌ Отмена")
        cancel.clicked.connect(self.reject)
        dlg_btns.addWidget(create)
        dlg_btns.addWidget(cancel)
        layout.addLayout(dlg_btns)
        applyTextFit(self)
        self.setMinimumSize(500, 450)
    
    def addColumn(self):
        dlg = AddColumnDialog(self, "")
        if dlg.exec():
            name, typ, default = dlg.getData()
            if name:
                col = {"name": name, "type": typ}
                self.columns.append(col)
                text = f"{name} ({typ})"
                if default:
                    text += f" [по умолч: {default}]"
                self.list.addItem(text)
    
    def removeColumn(self):
        r = self.list.currentRow()
        if r >= 0:
            self.list.takeItem(r)
            self.columns.pop(r)
    
    def getData(self):
        return self.name_edit.text().strip(), self.columns


class AddRecordDialog(QDialog):
    def __init__(self, parent, table, conn):
        super().__init__(parent)
        self.setWindowTitle("Добавить запись")
        self.setGeometry(300, 300, 400, 500)
        self.setFont(QFont("Arial", 10))
        
        self.table = table
        self.conn = conn
        self.entries = {}
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f"➕ Запись в '{table}'"))
        
        scroll = QScrollArea()
        widget = QWidget()
        scroll_layout = QVBoxLayout(widget)
        
        try:
            cursor = conn.cursor()
            cursor.execute(f"PRAGMA table_info('{table}')")
            for col in cursor.fetchall():
                name = col[1]
                typ = col[2]
                
                row = QHBoxLayout()
                row.addWidget(QLabel(f"{name} ({typ}):"))
                
                if typ.upper() == 'BOOLEAN':
                    entry = QComboBox()
                    entry.setFont(QFont("Arial", 10))
                    entry.addItems(["False", "True", "0", "1", "Нет", "Да"])
                else:
                    entry = QLineEdit()
                    entry.setFont(QFont("Arial", 10))
                
                self.entries[name] = (entry, typ)
                row.addWidget(entry)
                scroll_layout.addLayout(row)
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", str(e))
            self.reject()
        
        scroll.setWidget(widget)
        layout.addWidget(scroll)
        
        help = QLabel("Для BOOLEAN: True/1/Да или False/0/Нет")
        help.setStyleSheet("color: gray; font-size: 10px;")
        layout.addWidget(help)
        
        btns = QHBoxLayout()
        add = QPushButton("✅ Добавить")
        add.clicked.connect(self.accept)
        cancel = QPushButton("❌ Отмена")
        cancel.clicked.connect(self.reject)
        btns.addWidget(add)
        btns.addWidget(cancel)
        layout.addLayout(btns)
        applyTextFit(self)
        self.setMinimumSize(420, 450)
    
    def getValues(self):
        vals = []
        for name, (entry, typ) in self.entries.items():
            if isinstance(entry, QLineEdit):
                v = entry.text().strip()
            else:
                v = entry.currentText().strip()
            vals.append(None if v == "" else v)
        return vals


class ExportSettingsDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Настройки экспорта")
        self.setGeometry(300, 300, 400, 300)
        self.setFont(QFont("Arial", 10))
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("⚙️ Настройки фото"))
        
        self.include = QCheckBox("Включать фото в Excel")
        self.include.setChecked(True)
        layout.addWidget(self.include)
        
        self.save_files = QCheckBox("Сохранять фото отдельно")
        layout.addWidget(self.save_files)
        
        layout.addWidget(QLabel("Размер миниатюр:"))
        
        self.group = QButtonGroup(self)
        small = QRadioButton("Маленькие (80px)")
        medium = QRadioButton("Средние (100px)")
        large = QRadioButton("Большие (150px)")
        
        self.group.addButton(small, 80)
        self.group.addButton(medium, 100)
        self.group.addButton(large, 150)
        medium.setChecked(True)
        
        size = QHBoxLayout()
        size.addWidget(small)
        size.addWidget(medium)
        size.addWidget(large)
        layout.addLayout(size)
        
        btns = QHBoxLayout()
        ok = QPushButton("✅ Продолжить")
        ok.clicked.connect(self.accept)
        cancel = QPushButton("❌ Отмена")
        cancel.clicked.connect(self.reject)
        btns.addWidget(ok)
        btns.addWidget(cancel)
        layout.addLayout(btns)
        applyTextFit(self)
        self.setMinimumSize(380, 300)
    
    def getSettings(self):
        return {
            'include_images': self.include.isChecked(),
            'save_as_files': self.save_files.isChecked(),
            'image_size': self.group.checkedId()
        }


def main():
    try:
        app = QApplication(sys.argv)
        app.setStyle('Fusion')
        app.setStyleSheet(APP_STYLESHEET)
        
        font = QFont("Segoe UI", 9)
        font.setFamily("Segoe UI")
        font.setPointSize(9)
        app.setFont(font)
        
        window = ModernDatabaseApp()
        window.show()
        sys.exit(app.exec())
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
